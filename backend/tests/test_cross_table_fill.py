"""跨表数据填充模块测试（v0.6.0）。

测试覆盖 spec §Test Plan 全 7 节：
1. 上传 + 解析（multipart 双文件）
2. PATCH /config（主键 + 映射校验）
3. execute 匹配算法（exact / left / first / last / merge_multi / 内连接 / 空键值 / 大小写 / trim / overwrite / new_column / _filled 后缀 / 类型归一化 / 计数）
4. 双轨交付（preview ≤ 1000 + download_token / download 流式 xlsx）
5. 状态机 + 生命周期（pending → configured → executed / executed 锁 / 过期阻断 / DELETE CASCADE）
6. 列表 / 单查（status 过滤 + page + size + 默认 id 倒序）
7. SQL 日期函数不出现
"""
from __future__ import annotations

import inspect
import json
import re
import uuid
from io import BytesIO

import openpyxl
import pytest
from openpyxl import Workbook
from sqlalchemy import select, text, func
from starlette.datastructures import UploadFile as StarletteUploadFile

from app import models


# ---------- helpers ----------

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _make_xlsx(
    headers: list[str],
    rows: list[list],
    *,
    extra_row1_cells: dict[int, object] | None = None,
    duplicate_header_at: int | None = None,
    empty_workbook: bool = False,
) -> bytes:
    """手工构造一个最小 xlsx workbook。

    headers: row 1 表头文本
    rows: row 2+ 的二维数组
    extra_row1_cells: 额外 cell 写入 row 1（不影响表头）
    duplicate_header_at: 若指定，则把 row 1 的某个 cell 改成等于第 1 个表头以触发重名
    empty_workbook: 不写任何 sheet
    """
    wb = Workbook()
    if empty_workbook:
        # 默认会创建一个 active sheet，我们手动 remove 以触发「无 sheet」
        wb.remove(wb.active)
        buf = BytesIO()
        wb.save(buf)
        wb.close()
        buf.seek(0)
        return buf.getvalue()

    ws = wb.active
    ws.title = "Sheet1"
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    if duplicate_header_at is not None:
        # 把 duplicate_header_at 位置的 cell 改为第一列的 header
        ws.cell(row=1, column=duplicate_header_at, value=headers[0])
    if extra_row1_cells:
        for col, val in extra_row1_cells.items():
            ws.cell(row=1, column=col, value=val)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def _xlsx_upload(content: bytes, filename: str = "test.xlsx") -> tuple:
    """构造 (filename, bytes, content_type) 元组，等价于 httpx files 参数接收的形式。"""
    return (filename, content, XLSX_MIME)


async def _post_upload(client, target_bytes, base_bytes, *, expires_in_hours=None, target_name="t.xlsx", base_name="b.xlsx", target_ct=XLSX_MIME, base_ct=XLSX_MIME):
    """发送 POST /api/cross-table-fill/jobs；返回 response。"""
    data = {}
    if expires_in_hours is not None:
        data["expires_in_hours"] = str(expires_in_hours)
    return await client.post(
        "/api/cross-table-fill/jobs",
        data=data,
        files={
            "target_file": (target_name, target_bytes, target_ct),
            "base_file": (base_name, base_bytes, base_ct),
        },
    )


async def _patch_config(client, job_id, *, target_keys=None, base_keys=None, mappings=None, confirm_token=None, **kwargs):
    body = {
        "target_keys": target_keys or ["工号"],
        "base_keys": base_keys or ["EID"],
        "mappings": mappings or [{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
        "join_mode": kwargs.get("join_mode", "left"),
        "match_mode": kwargs.get("match_mode", "merge_multi"),
        "case_sensitive": kwargs.get("case_sensitive", True),
        "trim_strings": kwargs.get("trim_strings", True),
    }
    if confirm_token is not None:
        body["confirm_token"] = confirm_token
    elif "confirm_token" in body:
        pass
    return await client.patch(f"/api/cross-table-fill/jobs/{job_id}/config", json=body)


# ============================================================
# §1. 上传 + 解析
# ============================================================


async def test_upload_201_success(client, db):
    """TC01：双文件上传 → 201, headers + row_count + status=pending"""
    target = _make_xlsx(["工号", "姓名", "部门"], [["E001", "张三", None], ["E002", "李四", None]])
    base = _make_xlsx(["EID", "Department", "Email"], [["E001", "研发", "a@x.com"], ["E002", "测试", "b@x.com"]])
    r = await _post_upload(client, target, base, target_name="target.xlsx", base_name="base.xlsx")
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["status"] == "pending"
    assert body["target_headers"] == ["工号", "姓名", "部门"]
    assert body["base_headers"] == ["EID", "Department", "Email"]
    assert body["target_row_count"] == 2
    assert body["base_row_count"] == 2
    assert body["target_filename"] == "target.xlsx"
    assert body["base_filename"] == "base.xlsx"
    assert "job_id" in body
    # DB 落地校验
    count = db.execute(select(func.count()).select_from(models.CrossTableFillJob)).scalar()
    assert count == 1
    row_count = db.execute(select(func.count()).select_from(models.CrossTableFillRow)).scalar()
    assert row_count == 4  # 2 target + 2 base


async def test_upload_415_bad_mime(client):
    """TC02：MIME 非 .xlsx → 422"""
    target = _make_xlsx(["工号"], [["E001"]])
    base = _make_xlsx(["EID"], [["E001"]])
    r = await _post_upload(client, target, base, target_ct="text/plain", base_ct=XLSX_MIME)
    assert r.status_code in (422, 415), r.text


async def test_upload_413_too_large(client, monkeypatch):
    """TC03：文件 > 20 MB → 413"""
    # 暂时降低阈值以避免构造 20MB 文件
    from app.api import cross_table_fill as api_mod
    monkeypatch.setattr(api_mod, "MAX_BYTES", 1024)
    # 构造 ~2KB 文件
    target = _make_xlsx(["工号"], [["E001", "张三"] + [f"x{i}" for i in range(100)]])
    base = _make_xlsx(["EID"], [["E001"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 413, r.text


async def test_upload_400_malformed_xlsx(client):
    """TC04：malformed xlsx（MIME 是 .xlsx 但内部损坏）→ 400。

    注：openpyxl 自身拒绝保存 0-sheet 工作簿，故「真 xlsx 但无 sheet」不可构造。
    本测试覆盖「xlsx zip 内缺关键部件」的 broken file 场景。
    """
    import zipfile
    import io

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "[Content_Types].xml",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            b'<Default Extension="xml" ContentType="application/xml"/>'
            b'</Types>',
        )
        zf.writestr(
            "_rels/.rels",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            b'</Relationships>',
        )
        zf.writestr(
            "xl/workbook.xml",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            b'<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b'<sheets/>'
            b'</workbook>',
        )
    target = buf.getvalue()
    base = _make_xlsx(["EID"], [["E001"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 400, r.text


async def test_upload_422_duplicate_headers(client):
    """TC05：表头重复名 → 422"""
    target = _make_xlsx(["工号", "姓名", "姓名"], [["E001", "张三", "abc"]])
    base = _make_xlsx(["EID"], [["E001"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 422, r.text
    assert "duplicate" in r.text.lower() or "重复" in r.text or "name" in r.text.lower()


async def test_upload_422_empty_headers(client):
    """TC06：表头全空 → 422"""
    target = _make_xlsx(["", "  ", ""], [["v1", "v2", "v3"]])
    base = _make_xlsx(["EID"], [["E001"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 422, r.text


async def test_upload_400_bad_zip(client):
    """TC07：伪文件 → 400/422"""
    target = b"not a real xlsx file"
    base = _make_xlsx(["EID"], [["E001"]])
    r = await _post_upload(client, target, base)
    assert r.status_code in (400, 422), r.text


async def test_upload_headers_strip_and_order(client):
    """TC08：headers 含前后空格 → 自动 strip + 顺序保留"""
    target = _make_xlsx([" 工号 ", "  姓名", "部门 "], [["E001", "张三", "研发"]])
    base = _make_xlsx([" EID ", "Name"], [["E001", "Zhang"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["target_headers"] == ["工号", "姓名", "部门"]
    assert body["base_headers"] == ["EID", "Name"]


# ============================================================
# §2. PATCH /config
# ============================================================


async def _upload_and_get_job_id(client) -> int:
    """辅助：上传最小一对表返回 job_id。"""
    target = _make_xlsx(["工号", "姓名", "部门"], [["E001", "张三", None]])
    base = _make_xlsx(["EID", "Department"], [["E001", "研发"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 201, r.text
    return r.json()["job_id"]


async def test_config_200_success(client):
    """TC09：合法完整配置 → 200 + status=configured + digest"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "configured"
    digest = body["config_digest"]
    assert digest["mapping_count"] == 1
    assert digest["has_overwrite"] is False
    assert digest["has_new_column"] is True


async def test_config_422_target_keys_unknown_field(client):
    """TC10：target_keys 字段不在 target_headers → 422"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["不存在"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    assert r.status_code == 422, r.text


async def test_config_422_keys_length_mismatch(client):
    """TC11：target_keys 与 base_keys 长度不等 → 422"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["工号", "姓名"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    assert r.status_code == 422, r.text


async def test_config_422_mapping_base_unknown(client):
    """TC12：mapping.base_field 不在 base_headers → 422"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "不存在", "target_field": "部门", "mode": "new_column"}],
    )
    assert r.status_code == 422, r.text


async def test_config_422_mapping_target_unknown(client):
    """TC13：mapping.target_field 不在 target_headers → 422"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "不存在", "mode": "new_column"}],
    )
    assert r.status_code == 422, r.text


async def test_config_422_bad_mode(client):
    """TC14：mapping.mode 取值非法 → 422"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "append"}],
    )
    assert r.status_code == 422, r.text


async def test_config_409_overwrite_no_token(client):
    """TC15：mapping 含 overwrite 但 confirm_token 缺失 → 409"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "overwrite"}],
        confirm_token=None,
    )
    assert r.status_code == 409, r.text


async def test_config_409_new_column_with_token(client):
    """TC16：仅 new_column 模式但 confirm_token 非 None → 409"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
        confirm_token=str(uuid.uuid4()),
    )
    assert r.status_code == 409, r.text


async def test_config_200_overwrite_with_token(client):
    """配套 TC15：overwrite + 合法 token → 200"""
    job_id = await _upload_and_get_job_id(client)
    r = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "overwrite"}],
        confirm_token=str(uuid.uuid4()),
    )
    assert r.status_code == 200, r.text


async def test_config_warnings_empty_target_keys(client, db):
    """TC17：target_keys 在 target 表有空键值行 → warnings"""
    target = _make_xlsx(["工号", "姓名"], [[None, "张三"], [None, "李四"], ["E001", "王五"]])
    base = _make_xlsx(["EID", "Department"], [["E001", "研发"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 201, r.text
    job_id = r.json()["job_id"]
    r2 = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
    )
    assert r2.status_code == 200, r2.text
    body = r2.json()
    assert any("空" in w or "空键" in w or "空键值" in w for w in body["warnings"]), body["warnings"]


async def test_config_digest_reflects_options(client):
    """TC18：digest 字段正确反映 join_mode / match_mode / case_sensitive / trim_strings / has_overwrite / has_new_column"""
    target = _make_xlsx(["工号", "姓名", "部门"], [["E001", "张三", None]])
    base = _make_xlsx(["EID", "Department"], [["E001", "研发"]])
    r = await _post_upload(client, target, base)
    job_id = r.json()["job_id"]
    r2 = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[
            {"base_field": "Department", "target_field": "部门", "mode": "overwrite"},
        ],
        confirm_token=str(uuid.uuid4()),
        join_mode="inner",
        match_mode="first",
        case_sensitive=False,
        trim_strings=False,
    )
    assert r2.status_code == 200, r2.text
    digest = r2.json()["config_digest"]
    assert digest["has_overwrite"] is True
    assert digest["has_new_column"] is False
    assert digest["join_mode"] == "inner"
    assert digest["match_mode"] == "first"
    assert digest["case_sensitive"] is False
    assert digest["trim_strings"] is False


# ============================================================
# §3. execute 匹配算法
# ============================================================


async def _upload_pair(client, target_rows, base_rows, *, target_headers=None, base_headers=None):
    """辅助：上传两张表。target_rows/base_rows 为 list[list]，headers 必传。"""
    target = _make_xlsx(target_headers or ["工号", "姓名", "部门"], target_rows)
    base = _make_xlsx(base_headers or ["EID", "Department"], base_rows)
    r = await _post_upload(client, target, base)
    assert r.status_code == 201, r.text
    return r.json()["job_id"]


async def _configure_and_execute(
    client,
    job_id,
    *,
    mappings,
    join_mode="left",
    match_mode="merge_multi",
    case_sensitive=True,
    trim_strings=True,
):
    """辅助：PATCH config + POST execute，返回 execute response body."""
    has_overwrite = any(m["mode"] == "overwrite" for m in mappings)
    cfg_kwargs = dict(
        target_keys=["工号"], base_keys=["EID"],
        mappings=mappings,
        join_mode=join_mode,
        match_mode=match_mode,
        case_sensitive=case_sensitive,
        trim_strings=trim_strings,
    )
    if has_overwrite:
        cfg_kwargs["confirm_token"] = str(uuid.uuid4())
    r = await _patch_config(client, job_id, **cfg_kwargs)
    assert r.status_code == 200, r.text
    r2 = await client.post(f"/api/cross-table-fill/jobs/{job_id}/execute")
    assert r2.status_code == 200, r2.text
    return r2.json()


async def test_match_exact_hit(client):
    """TC20：exact 单行命中 → target 该行字段正确填充"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None]],
        base_rows=[["E001", "研发"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "overwrite"}],
    )
    assert body["summary"]["filled_count"] == 1
    assert body["summary"]["unmatched_count"] == 0
    assert body["summary"]["multi_match_count"] == 0
    assert body["preview"][0]["部门"] == "研发"


async def test_match_left_join_unmatched(client):
    """TC21：base 缺该 key → unmatched + 1, fill 保持原值"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None], ["E002", "李四", None]],
        base_rows=[["E001", "研发"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    assert body["summary"]["unmatched_count"] == 1
    assert body["summary"]["result_row_count"] == 2
    # 命中行：new_column 模式下末尾追加「部门_filled」列；未命中行留空
    assert body["preview"][0]["部门_filled"] == "研发"
    assert body["preview"][1]["部门_filled"] is None


async def test_match_first_mode_pick_first(client):
    """TC22：match_mode=first → candidates[0]"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None]],
        base_rows=[["E001", "研发"], ["E001", "测试"], ["E001", "运维"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
        match_mode="first",
    )
    assert body["preview"][0]["部门_filled"] == "研发"
    assert body["summary"]["multi_match_count"] == 1


async def test_match_last_mode_pick_last(client):
    """TC23：match_mode=last → candidates[-1]"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None]],
        base_rows=[["E001", "研发"], ["E001", "测试"], ["E001", "运维"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
        match_mode="last",
    )
    assert body["preview"][0]["部门_filled"] == "运维"


async def test_match_merge_multi_default(client):
    """TC24：base 3 行 → ';'.join + multi_match_count += 1"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None]],
        base_rows=[["E001", "研发"], ["E001", "测试"], ["E001", "运维"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    assert body["preview"][0]["部门_filled"] == "研发;测试;运维"
    assert body["summary"]["multi_match_count"] == 1


async def test_match_inner_join_drops_unmatched(client):
    """TC25：join_mode=inner → result_row_count 减少"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None], ["E002", "李四", None]],
        base_rows=[["E001", "研发"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
        join_mode="inner",
    )
    assert body["summary"]["result_row_count"] == 1
    assert body["summary"]["unmatched_count"] == 0  # inner 模式不计 unmatched


async def test_match_empty_target_key(client):
    """TC26：target 主键空 → unmatched + fill 不动"""
    job_id = await _upload_pair(
        client,
        target_rows=[[None, "张三", None], ["E001", "王五", None]],
        base_rows=[["E001", "研发"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    assert body["summary"]["unmatched_count"] == 1
    assert body["preview"][0]["部门_filled"] is None
    assert body["preview"][1]["部门_filled"] == "研发"


async def test_match_empty_base_key_excluded_from_index(client):
    """TC27：base 主键空 → 不入 index;该键的 target 行 unmatched"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None]],
        base_rows=[[None, "研发"], ["E001", "测试"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    # base E001 行存在 (测试) → 命中;base 空键行不入 index（不影响 target）
    assert body["preview"][0]["部门_filled"] == "测试"
    assert body["summary"]["unmatched_count"] == 0


async def test_match_case_insensitive(client):
    """TC28：case_sensitive=false → 大小写差异命中"""
    target = _make_xlsx(["工号", "姓名"], [["e001", "张三"]])
    base = _make_xlsx(["EID", "Department"], [["E001", "研发"]])
    r = await _post_upload(client, target, base)
    job_id = r.json()["job_id"]
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
        case_sensitive=False,
    )
    assert body["preview"][0]["姓名_filled"] == "研发"


async def test_match_trim_strings(client):
    """TC29：trim_strings=true → 空格差异命中"""
    target = _make_xlsx(["工号", "姓名"], [[" E001 ", "张三"]])
    base = _make_xlsx(["EID", "Department"], [["E001", "研发"]])
    r = await _post_upload(client, target, base)
    job_id = r.json()["job_id"]
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
        trim_strings=True,
    )
    assert body["preview"][0]["姓名_filled"] == "研发"


async def test_match_overwrite_mode(client):
    """TC30：overwrite 模式 → target 原列值被覆盖"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", "初始部门"]],
        base_rows=[["E001", "研发"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "overwrite"}],
    )
    assert body["preview"][0]["部门"] == "研发"
    assert body["summary"]["filled_count"] == 1


async def test_match_new_column_no_collision(client):
    """TC31：new_column 无冲突 → 末尾追加列"""
    target = _make_xlsx(
        ["工号", "姓名", "部门", "新部门"],  # 既有 target 已经有「新部门」列 → 触发 _filled
        [["E001", "张三", None, None]],
    )
    base = _make_xlsx(["EID", "Department"], [["E001", "研发"]])
    r = await _post_upload(client, target, base)
    job_id = r.json()["job_id"]
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "新部门", "mode": "new_column"}],
    )
    assert "新部门_filled" in body["preview_headers"]
    assert body["preview"][0]["新部门_filled"] == "研发"


async def test_match_new_column_with_collision(client):
    """TC32：new_column target_field 与 target 原列同名 → 加 _filled 后缀"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None]],
        base_rows=[["E001", "研发"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    assert "部门_filled" in body["preview_headers"]
    assert "部门" in body["preview_headers"]  # 原列还在
    assert body["preview"][0]["部门_filled"] == "研发"


async def test_match_new_column_multi_collision(client):
    """TC33：多次冲突 → _filled_2 / _filled_3"""
    # 让 target_headers 同时含「部门」「部门_filled」
    target = _make_xlsx(["工号", "部门", "部门_filled"], [["E001", None, None]])
    base = _make_xlsx(["EID", "Department1", "Department2"], [["E001", "研发", "测试"]])
    r = await _post_upload(client, target, base)
    job_id = r.json()["job_id"]
    rcfg = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[
            {"base_field": "Department1", "target_field": "部门", "mode": "new_column"},
            {"base_field": "Department2", "target_field": "部门", "mode": "new_column"},
        ],
    )
    assert rcfg.status_code == 200, rcfg.text
    r2 = await client.post(f"/api/cross-table-fill/jobs/{job_id}/execute")
    assert r2.status_code == 200, r2.text
    body = r2.json()
    # 第一个 mapping → 「部门_filled」; 第二个 mapping 又与「部门_filled」冲突 → 「部门_filled_2」
    assert "部门_filled" in body["preview_headers"]
    assert "部门_filled_2" in body["preview_headers"]


async def test_match_type_normalize_int_vs_str(client):
    """TC34：base 1(int) vs target "1"(str) → 命中"""
    target = _make_xlsx(["工号", "姓名"], [["1", "张三"]])
    base = _make_xlsx(["EID", "Department"], [[1, "研发"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 201, r.text
    job_id = r.json()["job_id"]
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
    )
    assert body["preview"][0]["姓名_filled"] == "研发"


async def test_match_type_normalize_strict_float_int(client):
    """TC35：base '1.0' (str) vs target '1' (str) → 不命中（视为不同）。

    注：openpyxl 把数字 cell `1.0` 与 `1` 一律读回为 int 1，无法构造纯数字的差异。
    本测试改用字符串 '1.0' vs '1' 验证主键归一化严格区分 ASCII 字符串。
    """
    target = _make_xlsx(["工号", "姓名"], [["1", "张三"]])
    base = _make_xlsx(["EID", "Department"], [["1.0", "研发"]])
    r = await _post_upload(client, target, base)
    assert r.status_code == 201, r.text
    job_id = r.json()["job_id"]
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
    )
    assert body["summary"]["unmatched_count"] == 1


async def test_match_multi_match_count(client):
    """TC36：2 target 行均命中 3 行 base → multi_match_count=2"""
    target = _make_xlsx(["工号", "姓名"], [["E001", "张三"], ["E002", "李四"]])
    base = _make_xlsx(["EID", "Department"], [
        ["E001", "研发"], ["E001", "测试"], ["E001", "运维"],
        ["E002", "市场"], ["E002", "行政"], ["E002", "财务"],
    ])
    r = await _post_upload(client, target, base)
    job_id = r.json()["job_id"]
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
    )
    assert body["summary"]["multi_match_count"] == 2


async def test_match_filled_count(client):
    """TC37：部分 mapping 全部为空 → 该行不计 filled；同时存在真 unmatched。

    数据布局：
    - target E001 / 张三 → base 有 E001 行但 Department 是空 → 匹配但 vals 空 → 不算 filled
    - target E002 / 李四 → base 有 E002 行且 Department='测试' → 命中且 fill → filled_count += 1
    - target E003 / 王五 → base 无此 key → unmatched += 1

    期望：filled_count = 1（仅 E002 算填充），unmatched_count = 1（仅 E003）。
    """
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None], ["E002", "李四", None], ["E003", "王五", None]],
        base_rows=[["E001", None], ["E002", "测试"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
    )
    assert body["summary"]["unmatched_count"] == 1
    assert body["summary"]["filled_count"] == 1


# ============================================================
# §4. 双轨交付
# ============================================================


async def test_execute_response_preview_and_token(client):
    """TC40：preview 长度 ≤ 1000 + download_token + download_url"""
    # 构造 1050 行 target
    target_rows = [[f"E{i:05d}", f"姓名{i}", None] for i in range(1050)]
    base_rows = [[f"E{i:05d}", "研发"] for i in range(1050)]
    job_id = await _upload_pair(
        client,
        target_rows=target_rows,
        base_rows=base_rows,
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
    )
    assert "preview" in body
    assert len(body["preview"]) <= 1000
    assert "download_token" in body
    assert len(body["download_token"]) >= 16
    assert body["download_url"].startswith("/api/cross-table-fill/jobs/")


async def test_download_with_token(client):
    """TC41：带正确 token → 200 + Content-Disposition + 文件可被 openpyxl 读回"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None]],
        base_rows=[["E001", "研发"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
    )
    token = body["download_token"]
    r = await client.get(f"/api/cross-table-fill/jobs/{job_id}/download?token={token}")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers.get("content-type", "")
    assert "attachment" in r.headers.get("content-disposition", "")
    # 解 xlsx 验证
    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # row 1 = headers; row 2 = data
    assert rows[0][0] == "工号"
    assert rows[1][0] == "E001"
    assert rows[1][1] == "张三"
    # new_column 模式下末尾应含「姓名_filled」
    assert "姓名_filled" in rows[0]


async def test_download_token_invalid(client):
    """TC42：token 不带 / 错误 → 401"""
    job_id = await _upload_pair(
        client,
        target_rows=[["E001", "张三", None]],
        base_rows=[["E001", "研发"]],
    )
    body = await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "姓名", "mode": "new_column"}],
    )
    token = body["download_token"]
    # 错误 token
    r1 = await client.get(f"/api/cross-table-fill/jobs/{job_id}/download?token=wrong-token")
    assert r1.status_code == 401, r1.text
    # 不带 token
    r2 = await client.get(f"/api/cross-table-fill/jobs/{job_id}/download")
    assert r2.status_code == 401, r2.text


async def test_download_status_not_executed(client):
    """TC43：status != executed → 409"""
    target = _make_xlsx(["工号", "姓名"], [["E001", "张三"]])
    base = _make_xlsx(["EID", "Department"], [["E001", "研发"]])
    r = await _post_upload(client, target, base)
    job_id = r.json()["job_id"]
    r2 = await client.get(f"/api/cross-table-fill/jobs/{job_id}/download?token=any")
    assert r2.status_code == 409, r2.text


# ============================================================
# §5. 状态机 + 生命周期
# ============================================================


async def test_status_machine_pending_to_executed(client, db):
    """TC50：pending → configured → executed"""
    job_id = await _upload_and_get_job_id(client)
    # pending
    job = db.get(models.CrossTableFillJob, job_id)
    assert job.status == "pending"
    # configured
    await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    db.refresh(job)
    assert job.status == "configured"
    # executed
    r = await client.post(f"/api/cross-table-fill/jobs/{job_id}/execute")
    assert r.status_code == 200, r.text
    db.refresh(job)
    assert job.status == "executed"


async def test_config_after_executed_blocked(client):
    """TC51：executed 后再 PATCH → 409"""
    job_id = await _upload_and_get_job_id(client)
    await _configure_and_execute(
        client, job_id,
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    r = await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    assert r.status_code == 409, r.text


async def test_expired_job_blocked(client, db, monkeypatch):
    """TC52：monkeypatch expires_at 为昨日 → blocked"""
    from datetime import date, timedelta
    job_id = await _upload_and_get_job_id(client)
    # 直接改 DB 把 expires_at 改为昨天
    job = db.get(models.CrossTableFillJob, job_id)
    job.expires_at = (date(2026, 8, 10) - timedelta(days=1)).isoformat()
    db.commit()
    # GET 仍能 GET 但 execute 应阻断
    r = await client.post(f"/api/cross-table-fill/jobs/{job_id}/execute")
    assert r.status_code == 409, r.text


async def test_delete_cascade(client, db):
    """TC53：DELETE 后 rows / configs 清空；CASCADE 验证"""
    job_id = await _upload_and_get_job_id(client)
    await _patch_config(
        client, job_id,
        target_keys=["工号"], base_keys=["EID"],
        mappings=[{"base_field": "Department", "target_field": "部门", "mode": "new_column"}],
    )
    # delete
    r = await client.delete(f"/api/cross-table-fill/jobs/{job_id}")
    assert r.status_code == 204, r.text
    # cascade 验证
    row_count = db.execute(
        select(func.count()).select_from(models.CrossTableFillRow).where(
            models.CrossTableFillRow.job_id == job_id
        )
    ).scalar()
    assert row_count == 0
    cfg_count = db.execute(
        select(func.count()).select_from(models.CrossTableFillConfig).where(
            models.CrossTableFillConfig.job_id == job_id
        )
    ).scalar()
    assert cfg_count == 0


async def test_get_job_404(client):
    """TC54：不存在 job_id → 404"""
    r = await client.get("/api/cross-table-fill/jobs/99999")
    assert r.status_code == 404, r.text


# ============================================================
# §6. 列表 / 单查
# ============================================================


async def test_list_jobs_pagination_and_filter(client, db):
    """TC60：status filter + page + size"""
    # 创建 3 个 job
    for i in range(3):
        target = _make_xlsx(["工号"], [[f"E{i:03d}"]])
        base = _make_xlsx(["EID"], [[f"E{i:03d}"]])
        r = await _post_upload(client, target, base)
        assert r.status_code == 201, r.text

    # 默认列表：3 条
    r = await client.get("/api/cross-table-fill/jobs")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total"] == 3
    assert body["page"] == 1
    assert body["size"] == 20

    # status filter
    r = await client.get("/api/cross-table-fill/jobs?status=pending")
    assert r.status_code == 200, r.text
    assert r.json()["total"] == 3

    # page + size
    r = await client.get("/api/cross-table-fill/jobs?page=1&size=2")
    assert r.status_code == 200, r.text
    assert len(r.json()["items"]) == 2
    assert r.json()["total"] == 3


async def test_list_jobs_default_order_by_id_desc(client):
    """TC61：默认按 id 倒序"""
    ids = []
    for i in range(2):
        target = _make_xlsx(["工号"], [[f"E{i:03d}"]])
        base = _make_xlsx(["EID"], [[f"E{i:03d}"]])
        r = await _post_upload(client, target, base)
        ids.append(r.json()["job_id"])

    r = await client.get("/api/cross-table-fill/jobs")
    items = r.json()["items"]
    assert [it["id"] for it in items] == sorted(ids, reverse=True)


# ============================================================
# §7. SQL 日期函数不出现
# ============================================================


async def test_no_sql_date_functions(client, db):
    """TC70：解析 / 查询 / 执行 SQL 文本不含 CURDATE / NOW / GETDATE"""
    # 该测试静态扫描源码
    from app import services, crud, api
    forbidden = ["CURDATE(", "NOW(", "CURRENT_DATE", "GETDATE(", "CURRENT_TIMESTAMP"]

    def _check(obj):
        src = inspect.getsource(obj)
        # 允许在 docstring 中出现字符但不出现函数调用形式
        for kw in forbidden:
            if kw in src:
                pytest.fail(f"{obj} contains forbidden SQL date function: {kw}")

    _check(services.cross_table_fill)
    _check(crud.cross_table_fill)
    _check(api.cross_table_fill)
