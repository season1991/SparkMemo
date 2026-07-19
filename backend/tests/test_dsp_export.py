"""DspUploadRow Excel 导出子模块测试（v0.5.8）。

测试覆盖（spec §Excel 导出子模块 §9.1）：
1. 200 正常路径：10 行事实 → 下载 → openpyxl 重新打开 → 内容与 JSON 端点一致
2. 200 空数据：0 行 batch → 表头 1 行
3. 404：id 不存在
4. 422 超限：200,001 行 → 中文 detail
5. 响应头：Content-Type + Content-Disposition 校验
6. 公式注入防护：country='=1+1' → 打开后 cell 值前缀 "'="

依赖：
- openpyxl 已存在（requirements.txt）
- pandas>=2.0 由本模块新增（services 层依赖）
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook


# 12 列固定表头（spec §4.1）
EXPECTED_HEADERS = [
    "ID", "上传批次ID", "国家", "类别", "配置代码", "配置名称",
    "数据类型", "TTL", "年月", "周编号", "周起始日", "数量",
]


# ---------- helpers ----------


def _parse_xlsx(content: bytes):
    """把响应 bytes 解析为 (headers, rows)。"""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)
    rows = list(rows_iter)
    wb.close()
    return list(headers), rows


def _sample_fact_rows() -> list[dict]:
    """构造 10 条事实行，覆盖各种 data_type / ttl / ym。"""
    return [
        {
            "country": "爱尔兰",
            "category": "交换机整机",
            "config_code": "X123",
            "config_name": "32Q-TOR-T3",
            "data_type": "Demand",
            "ttl": 4,
            "ym": "2026-07",
            "week": "WK27",
            "date": "2026-07-06",
            "quantity": 100,
        },
        {
            "country": "爱尔兰",
            "category": "交换机整机",
            "config_code": "X123",
            "config_name": "32Q-TOR-T3",
            "data_type": "Supply",
            "ttl": 4,
            "ym": "2026-07",
            "week": "WK28",
            "date": "2026-07-13",
            "quantity": 120,
        },
        {
            "country": "马来西亚",
            "category": "机箱",
            "config_code": "BD3300006913",
            "config_name": "8Q-TOR-T4",
            "data_type": "Demand",
            "ttl": None,
            "ym": "2026-07",
            "week": "WK27",
            "date": "2026-07-06",
            "quantity": 50,
        },
        {
            "country": "日本",
            "category": "电源",
            "config_code": "PWR-001",
            "config_name": "PSU-2000W",
            "data_type": "Demand",
            "ttl": 4,
            "ym": "2026-07",
            "week": "WK27",
            "date": "2026-07-06",
            "quantity": 25,
        },
        {
            "country": "日本",
            "category": "电源",
            "config_code": "PWR-001",
            "config_name": "PSU-2000W",
            "data_type": "Supply",
            "ttl": 4,
            "ym": "2026-08",
            "week": "WK31",
            "date": "2026-07-27",
            "quantity": 30,
        },
        {
            "country": None,
            "category": "线缆",
            "config_code": "CBL-001",
            "config_name": None,
            "data_type": "Demand",
            "ttl": 4,
            "ym": "2026-07",
            "week": "WK27",
            "date": "2026-07-06",
            "quantity": 10,
        },
        {
            "country": "爱尔兰",
            "category": "交换机整机",
            "config_code": "X456",
            "config_name": "64Q-TOR-T5",
            "data_type": "Demand",
            "ttl": 4,
            "ym": "2026-07",
            "week": "WK28",
            "date": "2026-07-13",
            "quantity": 80,
        },
        {
            "country": "马来西亚",
            "category": "机箱",
            "config_code": "BD3300006914",
            "config_name": "16Q-TOR-T6",
            "data_type": "Supply",
            "ttl": 4,
            "ym": "2026-07",
            "week": "WK28",
            "date": "2026-07-13",
            "quantity": 40,
        },
        {
            "country": "日本",
            "category": "电源",
            "config_code": "PWR-002",
            "config_name": "PSU-3000W",
            "data_type": "Demand",
            "ttl": 4,
            "ym": "2026-08",
            "week": "WK31",
            "date": "2026-07-27",
            "quantity": 15,
        },
        {
            "country": "爱尔兰",
            "category": "交换机整机",
            "config_code": "X789",
            "config_name": "128Q-TOR-T7",
            "data_type": "Demand",
            "ttl": 4,
            "ym": "2026-08",
            "week": "WK31",
            "date": "2026-07-27",
            "quantity": 5,
        },
    ]


# ---------- TC01~TC06 ----------


class TestRowsExport:
    """DspUploadRow Excel 导出端点。"""

    @pytest.mark.asyncio
    async def test_rows_export_200_basic(
        self, client, db, make_dsp_upload,
    ):
        """200 正常路径：10 行 → 下载 → 用 openpyxl 重新打开 → 断言列头 / 行数 / 内容。"""
        fact_rows = _sample_fact_rows()
        upload = make_dsp_upload(
            db,
            vendor="Arista",
            item="网络设备DSP横版",
            sub_item="机箱",
            version_date="2026-06-29",
            row_count=len(fact_rows),
            fact_rows=fact_rows,
        )

        resp = await client.get(f"/api/dsp-uploads/{upload.id}/rows/export")
        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        headers, rows = _parse_xlsx(resp.content)
        assert headers == EXPECTED_HEADERS
        assert len(rows) == 10

        # 抽查第 1 行：与 fact_rows[0] 字段一致
        # 列顺序：ID, upload_id, country, category, config_code, config_name,
        #        data_type, ttl, ym, week, date, quantity
        first_row = rows[0]
        assert first_row[1] == upload.id  # upload_id
        assert first_row[2] == "爱尔兰"  # country
        assert first_row[3] == "交换机整机"  # category
        assert first_row[4] == "X123"  # config_code
        assert first_row[5] == "32Q-TOR-T3"  # config_name
        assert first_row[6] == "Demand"  # data_type
        assert first_row[7] == 4  # ttl
        assert first_row[8] == "2026-07"  # ym
        assert first_row[9] == "WK27"  # week
        assert first_row[10] == "2026-07-06"  # date
        assert first_row[11] == 100  # quantity

    @pytest.mark.asyncio
    async def test_rows_export_200_empty(
        self, client, db, make_dsp_upload,
    ):
        """200 空数据：0 行 batch → sheet 仅表头 1 行（12 列）。"""
        upload = make_dsp_upload(
            db,
            vendor="Arista",
            item="X",
            sub_item="Y",
            version_date="2026-06-29",
            row_count=0,
        )

        resp = await client.get(f"/api/dsp-uploads/{upload.id}/rows/export")
        assert resp.status_code == 200

        headers, rows = _parse_xlsx(resp.content)
        assert headers == EXPECTED_HEADERS
        assert rows == []  # 仅表头，无数据行

    @pytest.mark.asyncio
    async def test_rows_export_404_unknown_id(self, client, db):
        """404：id 不存在。"""
        resp = await client.get("/api/dsp-uploads/999999/rows/export")
        assert resp.status_code == 404
        assert "not found" in resp.json()["detail"].lower()

    @pytest.mark.asyncio
    async def test_rows_export_422_over_limit(
        self, client, db, make_dsp_upload, monkeypatch,
    ):
        """422 超限：行数 > MAX_DSP_EXPORT_ROWS（默认 200_000）。"""
        from app.api import dsp_uploads as api_mod

        monkeypatch.setattr(api_mod, "MAX_DSP_EXPORT_ROWS", 5)

        fact_rows = _sample_fact_rows()  # 10 行
        upload = make_dsp_upload(
            db,
            vendor="Arista",
            item="X",
            sub_item="Y",
            version_date="2026-06-29",
            row_count=len(fact_rows),
            fact_rows=fact_rows,
        )

        resp = await client.get(f"/api/dsp-uploads/{upload.id}/rows/export")
        assert resp.status_code == 422
        detail = resp.json()["detail"]
        assert "超过上限" in detail or "exceeds" in detail.lower()

    @pytest.mark.asyncio
    async def test_rows_export_headers(
        self, client, db, make_dsp_upload,
    ):
        """响应头：Content-Type + Content-Disposition 含正确文件名。"""
        upload = make_dsp_upload(
            db, vendor="Arista", item="X", sub_item="Y",
            version_date="2026-06-29", row_count=0,
        )

        resp = await client.get(f"/api/dsp-uploads/{upload.id}/rows/export")
        assert resp.status_code == 200
        ct = resp.headers["content-type"]
        assert "spreadsheetml.sheet" in ct

        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        # 文件名匹配：dsp_upload_{id}_rows_{YYYYMMDD_HHMMSS}.xlsx
        assert f"dsp_upload_{upload.id}_rows_" in cd
        assert ".xlsx" in cd

    @pytest.mark.asyncio
    async def test_rows_export_formula_injection(
        self, client, db, make_dsp_upload,
    ):
        """公式注入防护：country='=1+1' → 打开后 cell 值前缀 "'="。"""
        fact_rows = [
            {
                "country": "=1+1",
                "category": "交换机整机",
                "config_code": "X123",
                "config_name": "32Q-TOR-T3",
                "data_type": "Demand",
                "ttl": 4,
                "ym": "2026-07",
                "week": "WK27",
                "date": "2026-07-06",
                "quantity": 100,
            },
        ]
        upload = make_dsp_upload(
            db, vendor="Arista", item="X", sub_item="Y",
            version_date="2026-06-29", row_count=1,
            fact_rows=fact_rows,
        )

        resp = await client.get(f"/api/dsp-uploads/{upload.id}/rows/export")
        assert resp.status_code == 200
        _, rows = _parse_xlsx(resp.content)
        # 列顺序：0=ID, 1=upload_id, 2=country
        assert rows[0][2] == "'=1+1"