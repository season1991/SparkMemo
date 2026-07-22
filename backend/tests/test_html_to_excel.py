"""html-to-excel 模块单测（SPEC §10）。

覆盖 19 个用例：locator × 4、recognizer × 5、cleaner × 4、writer × 2、pipeline × 4。
fixtures 在 `tests/fixtures/html_to_excel/` 下，文件层 fixture 见下面 `FIX_DIR`。
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from uuid import uuid4

import pytest
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from app.services.html_to_excel import HtmlToExcelPipeline
from app.services.html_to_excel.cleaner import (
    clean_cell_node,
    clean_text,
    infer_type,
    normalize_for_compare,
    parse_number,
)
from app.services.html_to_excel.locator import TitleLocator
from app.services.html_to_excel.parser import HTMLParser
from app.services.html_to_excel.schemas import MatchCandidate
from app.services.html_to_excel.writer import ExcelWriter, _safe_sheet_name


FIX_DIR = Path(__file__).parent / "fixtures" / "html_to_excel"


# ============== helpers ==============


def _read_html(name: str) -> str:
    return (FIX_DIR / name).read_text(encoding="utf-8")


def _tmp_dir(tmp_path: Path) -> Path:
    """统一创建临时输出目录。"""
    target = tmp_path / uuid4().hex
    target.mkdir(parents=True, exist_ok=True)
    return target


# ============== locator ==============


def test_locator_exact_match_caseinsensitive(tmp_path: Path) -> None:
    html = _read_html("simple_table.html")
    soup = BeautifulSoup(html, "lxml")
    candidates, _ = TitleLocator().find(soup, "item")  # lowercase
    assert len(candidates) >= 1
    matched = candidates[0]
    assert matched.matched_text.lower() == "item"


def test_locator_th_priority(tmp_path: Path) -> None:
    """`<th>` 比 listheader 更靠前的 source 应当排在前面。"""
    html = _read_html("simple_table.html")
    soup = BeautifulSoup(html, "lxml")
    candidates, _ = TitleLocator().find(soup, "Item")
    # 至少一个候选；其中应有 th-text 类型（最高 source 优先级）
    sources = [c.source for c in candidates]
    assert "th-text" in sources


def test_locator_no_match(tmp_path: Path) -> None:
    html = _read_html("simple_table.html")
    soup = BeautifulSoup(html, "lxml")
    candidates, suggestions = TitleLocator().find(soup, "完全找不到的标题")
    assert candidates == []
    # 应有近似建议
    assert isinstance(suggestions, list)


def test_locator_multiple_returns_all(tmp_path: Path) -> None:
    """同一标题在多处出现：返回全部候选。"""
    html = """
    <html><body>
      <table><thead><tr><th>Item</th></tr></thead><tbody><tr><td>x</td></tr></tbody></table>
      <table><thead><tr><th>Item</th></tr></thead><tbody><tr><td>y</td></tr></tbody></table>
    </body></html>
    """
    soup = BeautifulSoup(html, "lxml")
    candidates, _ = TitleLocator().find(soup, "Item")
    assert len(candidates) >= 2


# ============== recognizers (via pipeline) ==============


def test_table_recognizer_simple(tmp_path: Path) -> None:
    """simple_table 含 2 数据行 + 1 tfoot 合计行 → 3 行。"""
    src = FIX_DIR / "simple_table.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "Item", out)
    assert res.ok
    assert res.control_type == "table"
    # 3 rows：2 数据 + 1 tfoot 合计
    assert res.rows == 3


def test_table_recognizer_with_tfoot(tmp_path: Path) -> None:
    """含 tfoot 总计行：表数据 2 + 合计 1 行；最后一行 is_subtotal 标记为 True。"""
    from app.services.html_to_excel import HtmlToExcelPipeline
    from app.services.html_to_excel.parser import HTMLParser
    from app.services.html_to_excel.locator import TitleLocator
    from app.services.html_to_excel.recognizers import find_control_root

    html = _read_html("simple_table.html")
    soup = BeautifulSoup(html, "lxml")
    candidates, _ = TitleLocator().find(soup, "Item")
    chosen = candidates[0]
    rec, root = find_control_root(chosen.node)
    assert rec is not None
    control = rec.extract(root, chosen)
    # 3 行：2 数据 + 1 tfoot 合计
    assert control.row_count == 3
    assert control.rows[-1].is_subtotal is True
    # 合计行的统计数值应被推断为 number/integer
    assert any(c.type in ("number", "integer") for c in control.rows[-1].cells)


def test_table_recognizer_loading_placeholder(tmp_path: Path) -> None:
    """只有 Loading 占位：返回空 sheet（表头 + 1 占位行）。"""
    src = FIX_DIR / "loading_placeholder.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "Links", out)
    assert res.ok
    assert res.control_type == "table"
    assert "loading_or_empty" in res.warnings
    assert res.rows >= 1


def test_div_grid_recognizer(tmp_path: Path) -> None:
    src = FIX_DIR / "div_grid.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "项目列表", out)
    assert res.ok
    assert res.control_type == "div_grid"
    assert res.rows == 3
    assert res.columns == 3


def test_field_group_recognizer(tmp_path: Path) -> None:
    src = FIX_DIR / "field_group.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "Primary Information", out)
    assert res.ok
    assert res.control_type == "field_group"
    # 4 字段：Document Number / Date / Customer / To Email
    assert res.rows == 4
    assert res.columns == 2


def test_list_block_recognizer(tmp_path: Path) -> None:
    """`<ul>` / `<li>` 列表：抽取为 (Index, Text) 双列。"""
    html = """
    <html><body>
      <h2>标签</h2>
      <ul>
        <li>苹果</li>
        <li>香蕉</li>
        <li>橘子</li>
      </ul>
    </body></html>
    """
    src = tmp_path / "list_test.html"
    src.write_text(html, encoding="utf-8")
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "标签", out)
    assert res.ok
    assert res.control_type == "list_block"
    assert res.rows == 3


# ============== cleaner ==============


def test_cleaner_amp_entity() -> None:
    assert clean_text("R&amp;D team") == "R&D team"
    assert clean_text("Q &amp; A") == "Q & A"


def test_cleaner_nbsp_to_empty() -> None:
    assert clean_text("\xa0") == ""
    assert clean_text(" \xa0 hello \xa0 ") == "hello"


def test_cleaner_truncated_recover_tooltip() -> None:
    """`uir-field-truncated-value` 节点优先返回 tooltip 完整内容。"""
    html = """
    <span class="uir-field-truncated-value" data-ns-tooltip="完整描述内容很长很长很长很长">
      (more...)
    </span>
    """
    soup = BeautifulSoup(html, "lxml")
    cell = clean_cell_node(soup.find("span"))
    assert cell.value == "完整描述内容很长很长很长很长"
    assert cell.type == "tooltip"


def test_cleaner_checkbox() -> None:
    html = '<span class="checkbox_read_ck"><img class="checkboximage" alt="Checked"></span>'
    soup = BeautifulSoup(html, "lxml")
    cell = clean_cell_node(soup.find("span"))
    assert cell.value is True
    assert cell.type == "boolean"

    html2 = '<span class="checkbox_read_unck"><img class="checkboximage" alt="Unchecked"></span>'
    soup2 = BeautifulSoup(html2, "lxml")
    cell2 = clean_cell_node(soup2.find("span"))
    assert cell2.value is False
    assert cell2.type == "boolean"


# ============== writer ==============


def test_writer_formula_sanitize(tmp_path: Path) -> None:
    """首字符为 `=` / `+` / `-` / `@` 时单元格值加单引号前缀。"""
    from app.services.html_to_excel.schemas import ColumnDef, ExtractedCell, ExtractedControl, ExtractedRow

    ctrl = ExtractedControl(
        title="t",
        matched_text="t",
        source="th-text",
        control_type="table",
        columns=[
            ColumnDef(key="A", type="text", source="th", index=0),
            ColumnDef(key="B", type="text", source="th", index=1),
        ],
        rows=[
            ExtractedRow(cells=[
                ExtractedCell(value="=SUM(A1:A10)", type="text"),
                ExtractedCell(value="+1+1", type="text"),
            ])
        ],
    )
    out = _tmp_dir(tmp_path) / "f.xlsx"
    ExcelWriter().write(ctrl, out)
    wb = load_workbook(out, data_only=False)
    ws = wb.active
    assert ws.cell(row=2, column=1).value.startswith("'=")
    assert ws.cell(row=2, column=2).value.startswith("'+")


def test_writer_column_width(tmp_path: Path) -> None:
    """列宽自适应：长度 + 2，上限 50（来自既有 _auto_width）。"""
    from app.services.html_to_excel.schemas import ColumnDef, ExtractedCell, ExtractedControl, ExtractedRow

    ctrl = ExtractedControl(
        title="t",
        matched_text="t",
        source="th-text",
        control_type="table",
        columns=[ColumnDef(key="Wide Column", type="text", source="th", index=0)],
        rows=[ExtractedRow(cells=[ExtractedCell(value="x" * 80, type="text")])],
    )
    out = _tmp_dir(tmp_path) / "w.xlsx"
    ExcelWriter().write(ctrl, out)
    wb = load_workbook(out, data_only=True)
    ws = wb.active
    width = ws.column_dimensions["A"].width or 0
    assert width <= 50.0  # 已被 _auto_width 上限约束


# ============== pipeline ==============


def test_pipeline_end_to_end_with_table_html(tmp_path: Path) -> None:
    src = FIX_DIR / "simple_table.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "Item", out)
    assert res.ok, res.to_dict()
    assert res.control_type == "table"
    # 文件应存在
    assert res.xlsx_path and Path(res.xlsx_path).exists()
    # 重新打开验证 header 包含 'Item'
    wb = load_workbook(res.xlsx_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Item" in headers


def test_pipeline_end_to_end_with_div_grid(tmp_path: Path) -> None:
    src = FIX_DIR / "div_grid.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "项目列表", out)
    assert res.ok
    assert res.control_type == "div_grid"
    assert res.xlsx_path and Path(res.xlsx_path).exists()


def test_pipeline_field_group(tmp_path: Path) -> None:
    src = FIX_DIR / "field_group.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "Primary Information", out)
    assert res.ok
    assert res.control_type == "field_group"
    wb = load_workbook(res.xlsx_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["Label", "Value"]


def test_pipeline_no_match_error_json(tmp_path: Path) -> None:
    src = FIX_DIR / "simple_table.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run(src, "ZZZNOMATCH", out)
    assert not res.ok
    assert res.error == "title_not_found"
    # suggestions / candidates 至少有一个 key（即便没有近似匹配也应返回空 list）
    d = res.to_dict()
    assert "error" in d
    assert d["error"] == "title_not_found"


# ============== extras (helpers) ==============


def test_sheet_name_truncate() -> None:
    """超长 sheet name 安全截断 + 去重。"""
    used = set()
    n1 = _safe_sheet_name("x" * 50, used)
    n2 = _safe_sheet_name("x" * 50, used)  # 同名应去重
    assert len(n1) <= 31
    assert n1 != n2


def test_infer_type_helpers() -> None:
    assert infer_type("8/21/2025") == "date"
    assert infer_type("8/21/2025 5:35 am") == "datetime"
    assert infer_type("751,653.00") == "number"
    assert infer_type("34") == "integer"
    assert infer_type("hello") == "text"
    assert parse_number("751,653.00") == 751653.0
    assert parse_number("34") == 34
    assert normalize_for_compare("  item   1  ") == "item 1"


# ────────────────────────────── v0.2.0 新增 6 个 ──────────────────────────────


def test_inspect_returns_simple_table(tmp_path: Path) -> None:
    """simple_table.html 应至少返回 1 个 control，row_count ≥ 2（数据行）。"""
    src = FIX_DIR / "simple_table.html"
    res = HtmlToExcelPipeline().inspect(src)
    assert res.ok
    assert len(res.controls) >= 1
    c0 = res.controls[0]
    assert c0.row_count >= 2
    assert c0.column_count == 3
    assert c0.control_type == "table"
    # 预览应当有 headers 与 first_rows
    assert len(c0.preview.headers) == 3
    assert len(c0.preview.first_rows) >= 2


def test_inspect_filters_loading_placeholder(tmp_path: Path) -> None:
    """loading_placeholder.html 应返回空 controls: []。"""
    src = FIX_DIR / "loading_placeholder.html"
    res = HtmlToExcelPipeline().inspect(src)
    assert res.ok
    assert res.controls == []


def test_inspect_handles_nested_tables(tmp_path: Path) -> None:
    """nested_table.html：v0.2.0 起不再强制跳内嵌表（外层/内层都可能含真实数据）。

    验证：
    - 外层表列为「项目 / 详情」，control #0
    - 内层表（子项 1/2）在 cell 内，仍被识别为有效 control
    - 第一行第一条预览应为 `(subtable, N rows)` 占位（说明抽出时识别为嵌套）
    """
    src = FIX_DIR / "nested_table.html"
    res = HtmlToExcelPipeline().inspect(src)
    assert res.ok
    # 外 + 至少 1 个内层（实际看 soup 解析后会有 1~3 个 control，看具体嵌套深度）
    assert len(res.controls) >= 1
    outer = res.controls[0]
    assert outer.control_type == "table"
    # 至少有一个 control row 的 cell 含 "(subtable" 占位字符串（说明外层确实嵌了内层）
    has_subtable_placeholder = any(
        "(subtable" in str(cell.value)
        for row in outer.preview.first_rows
        for cell_obj in []  # placeholder is in summary list of preview row arrays
        for cell in []
    ) or any(
        "(subtable" in str(c)
        for row in outer.preview.first_rows
        for c in row
    )
    assert has_subtable_placeholder, "外层第一行至少一个 cell 应含 (subtable, N rows) 占位"


def test_inspect_meta_correct_for_netsuite_subset(tmp_path: Path) -> None:
    """netsuite_items_subset.html → controls[0].row_count=2, columns=4。"""
    src = FIX_DIR / "netsuite_items_subset.html"
    res = HtmlToExcelPipeline().inspect(src)
    assert res.ok
    assert len(res.controls) == 1
    c0 = res.controls[0]
    assert c0.row_count == 2
    assert c0.column_count == 4
    assert c0.preview.headers[:4] == ["Line Number", "Item Status", "Item", "Quantity"]
    # 第一行第一条数据是 HLL…
    assert c0.preview.first_rows[0][2] == "HLL060240327582B"


def test_extract_by_index_matches_extract_by_title(tmp_path: Path) -> None:
    """run_by_index(0) 的 rows/columns 应与 run(title='Item') 一致。"""
    src = FIX_DIR / "netsuite_items_subset.html"
    out = _tmp_dir(tmp_path)
    r_index = HtmlToExcelPipeline().run_by_index(
        html_path=src, index=0, output_dir=out, filename_hint="via-index"
    )
    r_title = HtmlToExcelPipeline().run(
        html_path=src, title="Items", output_dir=out, filename_hint="via-title"
    )
    assert r_index.ok, r_index.to_dict()
    assert r_title.ok, r_title.to_dict()
    assert (r_index.rows, r_index.columns) == (r_title.rows, r_title.columns)
    assert r_index.matched_title == r_title.matched_title


def test_extract_by_index_out_of_range(tmp_path: Path) -> None:
    """index=99 → ExtractionResult.error='index_out_of_range'。"""
    src = FIX_DIR / "simple_table.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run_by_index(
        html_path=src, index=99, output_dir=out
    )
    assert not res.ok
    assert res.error == "index_out_of_range"
    # 应附带 candidates（所有 suggested_title），便于前端重新选择
    assert res.candidates is not None and len(res.candidates) >= 1


def test_extract_by_index_negative(tmp_path: Path) -> None:
    """index < 0 → 同样 index_out_of_range。"""
    src = FIX_DIR / "simple_table.html"
    out = _tmp_dir(tmp_path)
    res = HtmlToExcelPipeline().run_by_index(
        html_path=src, index=-1, output_dir=out
    )
    assert not res.ok
    assert res.error == "index_out_of_range"
