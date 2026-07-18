"""DSP 上传模块测试（v0.5）。

测试覆盖 spec §Test Plan 全 7 节：
1. 文件名解析（纯函数）
2. Excel 解析（纯函数；含 R1/R2/C1/C2/C3/C4 + 数值容错）
3. POST /api/dsp-uploads（201 / 400 / 409 / 413 / 415 / 422）
4. GET 列表 / 详情 / 行分页
5. DELETE 级联
6. SQL 日期函数不出现
7. 真实文件回归
"""
from __future__ import annotations

import inspect
from datetime import date as _date
from io import BytesIO

import openpyxl
import pytest

from app.services.dsp_parser import (
    BadQuantityError,
    FactRow,
    SheetMissingError,
    parse_excel,
    parse_filename,
)


# ---------- helpers ----------

def _build_workbook(
    *,
    rows_data: list[dict],
    week_cols: list[tuple[str, str]],
    ym_map: dict[int, str] | None = None,
    sheet_name: str = "DSP",
) -> BytesIO:
    """手工构造一个 DSP 风格的最小 workbook。

    rows_data: 每项是一个 dict，可包含列号 -> 值的键值对。
        列号含义：4=Country, 5=Category, 6=ConfigCode, 10=DataType, 11=TTL。
            这些列号只是 v0.5.3 之前的硬编码位置占位；v0.5.3 起解析只依赖行 1 列头文本匹配。
        13..max_col 通过 week_cols 的索引对应。
    week_cols: [(week_str, date_str), ...]，从 col 13 起排列。
    ym_map: col -> 'YYYY-MM' 映射；如果给 None，会在 col 13 自动填 '2025-01' 整段。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Row 1：col 4..12 表头（不影响解析，仅占位）；col 13+ 携带 ym 标签
    ws.cell(row=1, column=4, value="*Country")
    ws.cell(row=1, column=5, value="Category")
    ws.cell(row=1, column=6, value="Config Code")
    ws.cell(row=1, column=10, value="Data Type")
    ws.cell(row=1, column=11, value="TTL")
    ws.cell(row=1, column=12, value="Update By")

    last_week_col = 13 + len(week_cols) - 1
    if ym_map is None:
        for i in range(len(week_cols)):
            ws.cell(row=1, column=13 + i, value="2025-01")
    else:
        for col, ym in ym_map.items():
            ws.cell(row=1, column=col, value=ym)

    # Row 2：周编号
    for i, (wk, _dt) in enumerate(week_cols):
        if wk:
            ws.cell(row=2, column=13 + i, value=wk)
    # Row 3：周起始日
    for i, (_wk, dt) in enumerate(week_cols):
        if dt:
            ws.cell(row=3, column=13 + i, value=dt)

    # Row 4+：数据
    for r_idx, row in enumerate(rows_data, start=4):
        for col, val in row.items():
            ws.cell(row=r_idx, column=col, value=val)

    out = BytesIO()
    wb.save(out)
    wb.close()
    out.seek(0)
    return out


def _build_custom_workbook(
    *,
    row1_cells: dict[int, str],
    row2_cells: dict[int, str] | None = None,
    row3_cells: dict[int, str] | None = None,
    data_rows: list[dict[int, object]] | None = None,
    sheet_name: str = "DSP",
) -> BytesIO:
    """v0.5.3 测试辅助：自由放置 row 1 cell 以验证列头匹配。

    row1_cells:    col -> header text（如 {4: '*Country', 7: 'Category', ...}）
    row2_cells / row3_cells: 同上
    data_rows:     每项一个 dict，col -> value（行号从 4 自动递增）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for c, v in row1_cells.items():
        ws.cell(row=1, column=c, value=v)
    for c, v in (row2_cells or {}).items():
        ws.cell(row=2, column=c, value=v)
    for c, v in (row3_cells or {}).items():
        ws.cell(row=3, column=c, value=v)
    for r_idx, row in enumerate(data_rows or [], start=4):
        for c, v in row.items():
            ws.cell(row=r_idx, column=c, value=v)

    out = BytesIO()
    wb.save(out)
    wb.close()
    out.seek(0)
    return out


def _xlsx_file(content: BytesIO, filename: str = "Arista-X-Y-061626.xlsx"):
    """构造 httpx 上传用的 (filename, bytes, content_type) 元组。"""
    return (filename, content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------- §1. 文件名解析 ----------

def test_parse_filename_arista():
    assert parse_filename("Arista-网络设备DSP横版-机箱-061626.xlsx") == (
        "Arista", "网络设备DSP横版", "机箱",
    )


def test_parse_filename_two_segments_raises():
    with pytest.raises(ValueError, match="at least 3 segments"):
        parse_filename("foo-bar.xlsx")


def test_parse_filename_no_extension_splits_on_dash():
    """没有扩展名 → 按整串 split('-')：2 段仍然 < 3，按通用规则抛 ValueError。"""
    with pytest.raises(ValueError, match="at least 3 segments"):
        parse_filename("no-extension")


def test_parse_filename_empty_raises():
    with pytest.raises(ValueError, match="filename is required"):
        parse_filename("")


def test_parse_filename_three_segments_ok():
    assert parse_filename("foo-bar-baz.xlsx") == ("foo", "bar", "baz")


def test_parse_filename_four_segments_drops_tail():
    """第 4 段及之后丢弃——这是 spec 文件名解析的设计。"""
    assert parse_filename("a-b-c-d.xlsx") == ("a", "b", "c")


# ---------- §2. Excel 解析 ----------

def test_parse_excel_basic_two_rows_three_weeks():
    """最小正向：1 个 Demand 行 × 3 个周列 = 3 条事实。"""
    rows = [{
        4: "Ireland", 5: "机箱", 6: "BD3300006913",
        10: "Demand", 11: 4,
        13: 1, 14: 2, 15: 3,
    }]
    weeks = [("WK01", "2024-12-30"), ("WK02", "2025-01-06"), ("WK03", "2025-01-13")]
    wb = _build_workbook(rows_data=rows, week_cols=weeks)
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 3
    assert all(fr.data_type == "Demand" for fr in facts)
    assert [fr.quantity for fr in facts] == [1, 2, 3]


def test_parse_excel_R1_country_and_config_both_empty_skips_row():
    """R1：Country 与 Config Code 同时空 → 整行跳过。"""
    rows = [
        {4: "", 5: "机箱", 6: "", 10: "Demand", 11: 4, 13: 5},  # 跳过
        {4: "Ireland", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 5},  # 保留
    ]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].config_code == "X"


def test_parse_excel_R1_only_country_empty_kept():
    """R1：只有 Country 空但 ConfigCode 非空 → 保留。"""
    rows = [{4: "", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].country is None
    assert facts[0].config_code == "X"


def test_parse_excel_R2_strict_demand_supply():
    """R2：data_type 不是 Demand/Supply → 整行跳过。"""
    rows = [
        {4: "X", 5: "机箱", 6: "X", 10: "GR", 11: 4, 13: 5},
        {4: "X", 5: "机箱", 6: "X", 10: "Demand PO", 11: 4, 13: 5},
        {4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 5},
        {4: "X", 5: "机箱", 6: "X", 10: "Supply", 11: 4, 13: 5},
        {4: "X", 5: "机箱", 6: "X", 10: "", 11: 4, 13: 5},
    ]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 2
    assert {fr.data_type for fr in facts} == {"Demand", "Supply"}


def test_parse_excel_R2_strips_whitespace_before_compare():
    """R2：Data Type 带尾空格 → strip 后等于 Demand，保留。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand ", 11: 4, 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1


def test_parse_excel_C1_week_number_empty_skips_column():
    """C1：row 2[col] 空 → 该 (行 × 列) 跳过；其它列对该行仍正常。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 1, 14: 2}]
    # week_cols[1] 给空字符串，模拟 week 缺失
    weeks = [("WK01", "2024-12-30"), ("", "2025-01-06")]
    wb = _build_workbook(rows_data=rows, week_cols=weeks)
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].week == "WK01"


def test_parse_excel_C2_date_empty_skips_column():
    """C2：row 3[col] 空 → 该 (行 × 列) 跳过。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 1, 14: 2}]
    weeks = [("WK01", "2024-12-30"), ("WK02", "")]
    wb = _build_workbook(rows_data=rows, week_cols=weeks)
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].week == "WK01"


def test_parse_excel_C4_quantity_zero_skipped():
    """C4：quantity=0 → 跳过。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 0}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert facts == []


def test_parse_excel_C4_quantity_none_and_empty_skipped():
    """C4：quantity=None 和 '' → 跳过。"""
    rows = [
        {4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: None},
        {4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: ""},
    ]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert facts == []


def test_parse_excel_quantity_nonnumeric_raises():
    """quantity 非数字字符串 → 抛 BadQuantityError（路由层转 400）。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: "abc"}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    with pytest.raises(BadQuantityError):
        parse_excel(wb.getvalue())


def test_parse_excel_quantity_non_integer_float_raises():
    """quantity 浮点非整数 → 抛 BadQuantityError。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 1.5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    with pytest.raises(BadQuantityError):
        parse_excel(wb.getvalue())


def test_parse_excel_quantity_string_int_accepted():
    """quantity 是数字字符串 → 入库 int。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: "42"}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].quantity == 42


def test_parse_excel_quantity_string_zero_skipped():
    """quantity 字符串 "0" → 跳过。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: "0"}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert facts == []


def test_parse_excel_ttl_invalid_string_becomes_none():
    """TTL 非整数字符串 → 入库 None，不阻断。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: "abc", 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].ttl is None


def test_parse_excel_ttl_float_integer_rounded():
    """TTL 浮点整数 4.0 → 入库 int(4)。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4.0, 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    facts = parse_excel(wb.getvalue())
    assert facts[0].ttl == 4
    assert isinstance(facts[0].ttl, int)


def test_parse_excel_ym_propagation_with_gap():
    """ym 前向传播：col 13='2025-01', col 14='', col 15='', col 16='2025-02'。
    col 13..15 入库 '2025-01'；col 16+ 入库 '2025-02'。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 1, 14: 2, 15: 3, 16: 4}]
    weeks = [
        ("WK01", "2024-12-30"),
        ("WK02", "2025-01-06"),
        ("WK03", "2025-01-13"),
        ("WK04", "2025-01-20"),
    ]
    ym_map = {13: "2025-01", 16: "2025-02"}
    wb = _build_workbook(rows_data=rows, week_cols=weeks, ym_map=ym_map)
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 4
    assert [fr.ym for fr in facts] == ["2025-01", "2025-01", "2025-01", "2025-02"]


def test_parse_excel_sheet_missing_raises():
    """sheet 名不是 DSP → SheetMissingError（路由层转 422）。"""
    wb = _build_workbook(rows_data=[], week_cols=[], sheet_name="Sheet1")
    with pytest.raises(SheetMissingError):
        parse_excel(wb.getvalue())


def test_parse_excel_no_valid_week_columns_returns_empty():
    """R3：完全没有有效周列（所有 row 2/3 都空）→ 0 条事实行。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4}]
    weeks = [("", "")]  # week 与 date 都空
    wb = _build_workbook(rows_data=rows, week_cols=weeks)
    facts = parse_excel(wb.getvalue())
    assert facts == []


# ---------- §3. POST /api/dsp-uploads ----------

async def test_post_upload_201_success(client, db):
    """最小正向：1 行 × 1 周列 → 1 条事实行。"""
    rows = [{4: "Ireland", 5: "机箱", 6: "BD3300006913",
             10: "Demand", 11: 4, 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])

    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_date": "2026-07-15",
        },
        files={"file": _xlsx_file(wb)},
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["vendor"] == "Arista"
    assert body["item"] == "X"
    assert body["sub_item"] == "Y"
    assert body["version_date"] == "2026-07-15"
    assert body["row_count"] == 1
    assert body["source_filename"] == "Arista-X-Y-061626.xlsx"
    assert body["created_at"] == _date.today().isoformat()  # spec: Python 传入今天

    # DB 落地校验
    from sqlalchemy import select, func
    from app import models
    row_count = db.execute(
        select(func.count()).select_from(models.DspUploadRow)
    ).scalar()
    assert row_count == 1
    upload_count = db.execute(
        select(func.count()).select_from(models.DspUpload)
    ).scalar()
    assert upload_count == 1


async def test_post_upload_400_version_date_bad_format(client):
    wb = _build_workbook(rows_data=[], week_cols=[])
    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_date": "2026/07/15",
        },
        files={"file": _xlsx_file(wb)},
    )
    assert response.status_code == 400
    assert "YYYY-MM-DD" in response.json()["detail"]


async def test_post_upload_422_missing_form_field(client):
    """v0.5.1：四个 Form 字段全部必填，缺一个由 FastAPI 自动 422。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    # 故意漏掉 sub_item
    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "version_date": "2026-07-15",
        },
        files={"file": _xlsx_file(wb)},
    )
    assert response.status_code == 422


async def test_post_upload_accepts_user_edited_vendor_item_sub_item(client, db):
    """v0.5.1：用户在前端编辑后的 vendor/item/sub_item 会落地，与文件名解析路径无关。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    # 文件名是 Arista-X-Y-061626.xlsx，但上传时改 sub_item 为 "Z"
    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Z",       # 故意改
            "version_date": "2026-07-15",
        },
        files={"file": _xlsx_file(wb)},
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["vendor"] == "Arista"
    assert body["item"] == "X"
    assert body["sub_item"] == "Z"        # 后端以 Form 传入值为准
    assert body["source_filename"] == "Arista-X-Y-061626.xlsx"  # 文件名仅审计


async def test_post_upload_400_quantity_nonnumeric(client):
    """quantity 非数字 → 400。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: "abc"}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_date": "2026-07-15",
        },
        files={"file": _xlsx_file(wb)},
    )
    assert response.status_code == 400
    assert "non-numeric" in response.json()["detail"]


async def test_post_upload_415_wrong_mime(client):
    wb = _build_workbook(rows_data=[], week_cols=[])
    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_date": "2026-07-15",
        },
        files={"file": ("Arista-X-Y-061626.xlsx", wb.getvalue(), "text/plain")},
    )
    assert response.status_code == 415


async def test_post_upload_413_oversize(monkeypatch, client):
    """文件 > 20 MB → 413（用 monkeypatch 改小阈值，加快测试）。"""
    from app.api import dsp_uploads
    monkeypatch.setattr(dsp_uploads, "MAX_BYTES", 1024)
    big = b"x" * 2048
    wb = openpyxl.Workbook()
    wb.active.title = "DSP"
    bio = BytesIO()
    wb.save(bio)
    wb.close()
    bio.seek(0)
    # 用大占位 payload + 合法 .xlsx 文件名 + MIME，强制触发 size 校验
    payload_bytes = bio.getvalue() + b"\x00" * 1024  # 超过 1024 阈值
    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_date": "2026-07-15",
        },
        files={"file": ("Arista-X-Y-061626.xlsx", payload_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 413


async def test_post_upload_409_duplicate_version(client):
    """同 (vendor, item, sub_item, version_date) 重传 → 409，detail 含 upload_id。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    payload = {
        "vendor": "Arista",
        "item": "X",
        "sub_item": "Y",
        "version_date": "2026-07-15",
    }

    r1 = await client.post(
        "/api/dsp-uploads",
        data=payload,
        files={"file": _xlsx_file(wb)},
    )
    assert r1.status_code == 201
    first_id = r1.json()["id"]

    r2 = await client.post(
        "/api/dsp-uploads",
        data=payload,
        files={"file": _xlsx_file(wb)},
    )
    assert r2.status_code == 409
    detail = r2.json()["detail"]
    assert f"upload_id={first_id}" in detail


async def test_post_upload_422_wrong_sheet(client):
    """sheet 名不是 DSP → 422。"""
    wb = _build_workbook(rows_data=[], week_cols=[], sheet_name="Sheet1")
    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_date": "2026-07-15",
        },
        files={"file": _xlsx_file(wb)},
    )
    assert response.status_code == 422


async def test_post_upload_then_resubmit_after_delete_succeeds(client):
    """删除后可以重传同版本（验证 409 仅在存在时阻断）。"""
    rows = [{4: "X", 5: "机箱", 6: "X", 10: "Demand", 11: 4, 13: 5}]
    wb = _build_workbook(rows_data=rows, week_cols=[("WK01", "2024-12-30")])
    payload = {
        "vendor": "Arista",
        "item": "X",
        "sub_item": "Y",
        "version_date": "2026-07-15",
    }

    r1 = await client.post(
        "/api/dsp-uploads",
        data=payload,
        files={"file": _xlsx_file(wb)},
    )
    assert r1.status_code == 201
    upload_id = r1.json()["id"]

    d = await client.delete(f"/api/dsp-uploads/{upload_id}")
    assert d.status_code == 204

    r2 = await client.post(
        "/api/dsp-uploads",
        data=payload,
        files={"file": _xlsx_file(wb)},
    )
    assert r2.status_code == 201
    # 二次插入的 version_date 相同即可（id 可能复用，autoincrement 行为依赖方言）
    assert r2.json()["version_date"] == "2026-07-15"
    assert r2.json()["vendor"] == "Arista"


# ---------- §4. GET 列表 / 详情 / 行分页 ----------

async def test_get_list_default_desc_by_id(client, make_dsp_upload, db):
    a = make_dsp_upload(db, version_date="2026-01-01")
    b = make_dsp_upload(db, version_date="2026-02-01")
    c = make_dsp_upload(db, version_date="2026-03-01")

    response = await client.get("/api/dsp-uploads")
    assert response.status_code == 200
    body = response.json()
    assert [it["id"] for it in body["items"]] == [c.id, b.id, a.id]
    assert body["total"] == 3
    assert body["page"] == 1
    assert body["size"] == 20


async def test_get_list_pagination(client, make_dsp_upload, db):
    for i in range(25):
        make_dsp_upload(db, version_date=f"2026-01-{i + 1:02d}")

    response = await client.get("/api/dsp-uploads", params={"page": 2, "size": 10})
    assert response.status_code == 200
    body = response.json()
    assert len(body["items"]) == 10
    assert body["total"] == 25


async def test_get_detail(client, make_dsp_upload, db):
    upload = make_dsp_upload(db, row_count=3)
    response = await client.get(f"/api/dsp-uploads/{upload.id}")
    assert response.status_code == 200
    assert response.json()["id"] == upload.id


async def test_get_detail_404(client):
    response = await client.get("/api/dsp-uploads/999999")
    assert response.status_code == 404


async def test_get_rows_pagination(client, make_dsp_upload, db):
    fact_rows = [
        {
            "country": f"C{i}", "category": "机箱", "config_code": "X",
            "data_type": "Demand", "ttl": 4,
            "ym": "2025-01", "week": f"WK{i + 1:02d}",
            "date": "2024-12-30", "quantity": i + 1,
        }
        for i in range(25)
    ]
    upload = make_dsp_upload(db, row_count=25, fact_rows=fact_rows)

    response = await client.get(
        f"/api/dsp-uploads/{upload.id}/rows", params={"page": 2, "size": 10},
    )
    assert response.status_code == 200
    body = response.json()
    assert len(body["items"]) == 10
    assert body["total"] == 25
    # 第二页第一条 id 应为 11（按 id 升序）
    assert body["items"][0]["id"] == 11


async def test_get_rows_404_when_upload_missing(client):
    response = await client.get("/api/dsp-uploads/999999/rows")
    assert response.status_code == 404


# ---------- §5. DELETE 级联 ----------

async def test_delete_cascades_to_rows(client, make_dsp_upload, db):
    from sqlalchemy import select, func
    from app import models

    fact_rows = [{
        "country": "X", "category": "机箱", "config_code": "X",
        "data_type": "Demand", "ttl": 4,
        "ym": "2025-01", "week": "WK01",
        "date": "2024-12-30", "quantity": 1,
    }]
    upload = make_dsp_upload(db, row_count=1, fact_rows=fact_rows)

    # sanity
    pre = db.execute(
        select(func.count()).select_from(models.DspUploadRow)
        .where(models.DspUploadRow.upload_id == upload.id)
    ).scalar()
    assert pre == 1

    response = await client.delete(f"/api/dsp-uploads/{upload.id}")
    assert response.status_code == 204

    after = db.execute(
        select(func.count()).select_from(models.DspUploadRow)
        .where(models.DspUploadRow.upload_id == upload.id)
    ).scalar()
    assert after == 0

    detail = await client.get(f"/api/dsp-uploads/{upload.id}")
    assert detail.status_code == 404


async def test_delete_404(client):
    response = await client.delete("/api/dsp-uploads/999999")
    assert response.status_code == 404


# ---------- §6. SQL 日期函数不出现 ----------

def test_no_sql_date_functions_in_module_source():
    """dsp_parser / crud.dsp_upload / api.dsp_uploads 的源文本中不出现数据库日期函数。"""
    from app.services import dsp_parser
    from app.crud import dsp_upload as crud_mod
    from app.api import dsp_uploads as api_mod

    forbidden = ("CURDATE(", "NOW()", "CURRENT_DATE", "GETDATE(")
    for mod in (dsp_parser, crud_mod, api_mod):
        src = inspect.getsource(mod)
        for token in forbidden:
            assert token not in src, f"{mod.__name__} contains forbidden token {token!r}"


# ---------- §7. 真实文件回归 ----------

REAL_FILE = "tests/fixtures/Arista-网络设备DSP横版-机箱-061626.xlsx"


async def test_real_file_regression(client):
    """真实样本文件：Demand + Supply 行展开后应为 366 条事实行。"""
    with open(REAL_FILE, "rb") as f:
        content = f.read()

    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "网络设备DSP横版",
            "sub_item": "机箱",
            "version_date": "2026-07-15",
        },
        files={"file": (
            "Arista-网络设备DSP横版-机箱-061626.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )},
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["vendor"] == "Arista"
    assert body["item"] == "网络设备DSP横版"
    assert body["sub_item"] == "机箱"
    assert body["row_count"] == 366
    assert body["created_at"] == _date.today().isoformat()

    # 抽查事实行：data_type 必须只有 Demand/Supply
    rows_resp = await client.get(
        f"/api/dsp-uploads/{body['id']}/rows", params={"page": 1, "size": 1000},
    )
    assert rows_resp.status_code == 200
    items = rows_resp.json()["items"]
    assert rows_resp.json()["total"] == 366
    assert len(items) == 366
    assert {it["data_type"] for it in items} <= {"Demand", "Supply"}

    # 任取一行抽查：所有字段非空且类型正确
    sample = items[0]
    assert isinstance(sample["id"], int)
    assert isinstance(sample["country"], (str, type(None)))
    assert isinstance(sample["category"], (str, type(None)))
    assert isinstance(sample["config_code"], (str, type(None)))
    assert sample["data_type"] in ("Demand", "Supply")
    assert isinstance(sample["ttl"], (int, type(None)))
    assert len(sample["ym"]) == 7 and sample["ym"][4] == "-"
    assert sample["week"].startswith("WK")
    assert len(sample["date"]) == 10 and sample["date"][4] == "-" and sample["date"][7] == "-"
    assert isinstance(sample["quantity"], int) and sample["quantity"] > 0


# ---------- v0.5.3 列头文本匹配测试 ----------

def _sample_facts_row():
    """v0.5.3 测试通用数据行（col → value）。"""
    return {
        2: "Ireland",       # country
        3: "机箱",          # category
        4: "BD3300006913",  # config_code
        9: "Demand",        # data_type
        10: 4,              # ttl
        # cols 12/13 写 ym='2025-01'；分别两个周列 14, 15
    }


def test_parse_excel_v0_5_3_reordered_columns_layout_a():
    """Layout A：列头文本在最 "原始" 的位置上（col 2..6 + col 9..10），与 _build_workbook 等价。
    列头匹配应正常解析。"""
    wb = _build_custom_workbook(
        row1_cells={
            2: "*Country",
            3: "Category",
            4: "Config Code",
            9: "Data Type",
            10: "TTL",
            12: "Update By",
            14: "2025-01",
        },
        row2_cells={14: "WK01"},
        row3_cells={14: "2024-12-30"},
        data_rows=[
            {2: "Ireland", 3: "机箱", 4: "BD3300006913", 9: "Demand", 10: 4, 14: 5},
        ],
    )
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    f = facts[0]
    assert f.country == "Ireland"
    assert f.category == "机箱"
    assert f.config_code == "BD3300006913"
    assert f.data_type == "Demand"
    assert f.ttl == 4
    assert f.week == "WK01"
    assert f.date == "2024-12-30"
    assert f.quantity == 5
    assert f.ym == "2025-01"


def test_parse_excel_v0_5_3_reordered_columns_layout_b():
    """Layout B：列头文本移到完全不同的位置上（*Country 在 col 6、TTL 在 col 4、DataType 在 col 12）。
    验证列头匹配自适应列重排。"""
    wb = _build_custom_workbook(
        row1_cells={
            4: "TTL",                      # 数据在物理位置 4
            6: "*Country",                 # Country 在 col 6
            7: "Category",                 # Category 在 col 7
            9: "Config Code",              # config_code 在 col 9
            12: "Data Type",                # data_type 在 col 12
            14: "2025-01",
        },
        row2_cells={14: "WK01"},
        row3_cells={14: "2024-12-30"},
        data_rows=[
            {4: 7, 6: "Ireland", 7: "机箱", 9: "BD3300006913", 12: "Demand", 14: 8},
        ],
    )
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    f = facts[0]
    assert f.country == "Ireland"   # 来自 col 6
    assert f.category == "机箱"    # 来自 col 7
    assert f.config_code == "BD3300006913"  # 来自 col 9
    assert f.data_type == "Demand"  # 来自 col 12
    assert f.ttl == 7              # 来自 col 4
    assert f.quantity == 8


def test_parse_excel_v0_5_3_strips_asterisk_and_whitespace():
    """行 1 列头文本 '  *Country  '  与 '*Country' 与 'country' 等价。"""
    wb = _build_custom_workbook(
        row1_cells={
            2: " *Country ",
            3: "Category",
            4: "Config Code",
            9: "Data Type",
            10: "TTL",
            12: "2025-01",
        },
        row2_cells={12: "WK01"},
        row3_cells={12: "2024-12-30"},
        data_rows=[
            {2: "Ireland", 3: "机箱", 4: "X", 9: "Demand", 10: 4, 12: 5},
        ],
    )
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].country == "Ireland"


def test_parse_excel_v0_5_3_case_insensitive():
    """列头文本大小写不敏感：'country' / 'COUNTRY' / 'Country' 都匹配。"""
    wb = _build_custom_workbook(
        row1_cells={
            2: "country",
            3: "CATEGORY",
            4: "config code",
            9: "DATA TYPE",
            10: "ttl",
            12: "2025-01",
        },
        row2_cells={12: "WK01"},
        row3_cells={12: "2024-12-30"},
        data_rows=[
            {2: "Ireland", 3: "机箱", 4: "X", 9: "Demand", 10: 4, 12: 5},
        ],
    )
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].data_type == "Demand"


def test_parse_excel_v0_5_3_aliases():
    """config_code 别名：'Config Code' / 'ConfigCode' / 'configcode' 都匹配。
    data_type 别名：'Data Type' / 'DataType' / 'datatype' 都匹配。"""
    wb = _build_custom_workbook(
        row1_cells={
            2: "Country",
            3: "Category",
            4: "ConfigCode",          # 无空格的别名
            9: "DataType",            # 无空格的别名
            10: "TTL",
            12: "2025-01",
        },
        row2_cells={12: "WK01"},
        row3_cells={12: "2024-12-30"},
        data_rows=[
            {2: "Ireland", 3: "机箱", 4: "X", 9: "Demand", 10: 4, 12: 5},
        ],
    )
    facts = parse_excel(wb.getvalue())
    assert len(facts) == 1
    assert facts[0].config_code == "X"
    assert facts[0].data_type == "Demand"


def test_parse_excel_v0_5_3_missing_country_raises_BadHeaderError():
    """行 1 缺 country 列 → BadHeaderError。"""
    wb = _build_custom_workbook(
        row1_cells={
            # 故意没有 Country 列
            3: "Category",
            4: "Config Code",
            9: "Data Type",
            10: "TTL",
            12: "2025-01",
        },
        row2_cells={12: "WK01"},
        row3_cells={12: "2024-12-30"},
        data_rows=[
            {3: "机箱", 4: "X", 9: "Demand", 10: 4, 12: 5},
        ],
    )
    from app.services.dsp_parser import BadHeaderError
    with pytest.raises(BadHeaderError) as ei:
        parse_excel(wb.getvalue())
    assert "country" in str(ei.value)


async def test_post_upload_422_missing_key_column(client):
    """POST 上传缺关键列的 workbook → 422 + detail 提示列名。"""
    wb = _build_custom_workbook(
        row1_cells={
            2: "Country",
            3: "Category",
            4: "Config Code",
            # 故意缺 data_type
            10: "TTL",
            12: "2025-01",
        },
        row2_cells={12: "WK01"},
        row3_cells={12: "2024-12-30"},
        data_rows=[
            {2: "Ireland", 3: "机箱", 4: "X", 10: 4, 12: 5},
        ],
    )
    response = await client.post(
        "/api/dsp-uploads",
        data={
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_date": "2026-07-15",
        },
        files={"file": _xlsx_file(wb)},
    )
    assert response.status_code == 422, response.text
    detail = response.json()["detail"]
    assert "data_type" in detail