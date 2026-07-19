"""透视查询 Excel 导出子模块测试（v0.5.8）。

测试覆盖（spec §Excel 导出子模块 §9.2）：
1. 200 demand 模式：sheet 1 列头 = [7 列基础] + period_columns
2. 200 demand_plus_supply 模式：含 TTL_GAP / Rolling_TTLGAP 行
3. 200 sheet 2 快照：5 列版本日期 `; ` 拼接
4. 422 级联校验失败
5. 422 笛卡尔积超限（monkeypatch MAX_CARTESIAN）
6. 422 demand_plus_supply + 多 version_date
7. 200 空 row_groups：sheet 1 仅表头，sheet 2 仍正确
"""
from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.crud import pivot_query


# ---------- helpers ----------


@pytest.fixture
def make_week_dt():
    """复用 test_pivot_query.py 的 week_dt fixture 模式。"""
    from sqlalchemy import text

    from app import models

    def _ensure_table(db):
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS week_dt (
                dt VARCHAR(10) PRIMARY KEY,
                year_id SMALLINT NOT NULL,
                month_id TINYINT NOT NULL,
                week_id TINYINT NOT NULL,
                is_week_start BOOLEAN NOT NULL DEFAULT 0
            )
        """))
        db.commit()

    def factory(db, dt, *, year_id, month_id, week_id, is_week_start=False):
        _ensure_table(db)
        row = models.WeekDt(
            dt=dt,
            year_id=year_id,
            month_id=month_id,
            week_id=week_id,
            is_week_start=is_week_start,
        )
        db.add(row)
        db.commit()
        return row

    return factory


def _parse_xlsx_all_sheets(content: bytes) -> dict[str, tuple[list, list]]:
    """返回 {sheet_name: (headers, rows)}。"""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    result = {}
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            result[ws.title] = ([], [])
            continue
        rows = list(rows_iter)
        result[ws.title] = (list(headers), rows)
    wb.close()
    return result


# ---------- TC01: demand 模式 ----------


class TestPivotExportDemand:
    """pivot_type='demand' 导出。"""

    @pytest.mark.asyncio
    async def test_pivot_export_200_demand(
        self, client, db, make_dsp_upload, make_week_dt,
    ):
        fact_rows = [
            {
                "country": "爱尔兰",
                "category": "交换机整机",
                "config_code": "X123",
                "config_name": "32Q-TOR-T3",
                "data_type": "Demand",
                "ttl": 4,
                "ym": "2026-07",
                "week": "WK27",
                "date": "2026-07-06",
                "quantity": 100,
            },
        ]
        make_dsp_upload(
            db, vendor="Arista", item="X", sub_item="Y",
            version_date="2026-06-29", fact_rows=fact_rows,
        )
        make_week_dt(db, "2026-07-06", year_id=2026, month_id=7,
                     week_id=27, is_week_start=True)
        make_week_dt(db, "2026-07-13", year_id=2026, month_id=7,
                     week_id=28, is_week_start=True)

        payload = {
            "pivot_type": "demand",
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_dates": ["2026-06-29"],
            "years": [2026],
        }
        resp = await client.post("/api/pivot-query/export", json=payload)
        assert resp.status_code == 200

        sheets = _parse_xlsx_all_sheets(resp.content)
        # sheet 1
        assert "透视结果" in sheets
        headers, rows = sheets["透视结果"]
        # 列头：国家 / 类别 / 配置代码 / 配置名称 / 数据类型 / TTL / 版本日期 /
        #       2026-07-06 / 2026-07-13
        assert headers[0] == "国家"
        assert headers[4] == "数据类型"
        assert headers[6] == "版本日期"
        assert headers[7:] == ["2026-07-06", "2026-07-13"]
        assert len(rows) == 1
        assert rows[0][0] == "爱尔兰"
        assert rows[0][4] == "Demand"
        assert rows[0][6] == "2026-06-29"
        # 2026-07-06 有数据 100；2026-07-13 COALESCE 兜底 0
        assert rows[0][7] == 100
        assert rows[0][8] == 0


# ---------- TC02: demand_plus_supply 模式 ----------


class TestPivotExportDemandPlusSupply:
    """pivot_type='demand_plus_supply' 导出：4 行/组 + sheet 2 快照。"""

    @pytest.mark.asyncio
    async def test_pivot_export_200_demand_plus_supply(
        self, client, db, make_dsp_upload, make_week_dt,
    ):
        fact_rows = [
            {
                "country": "爱尔兰", "category": "交换机整机",
                "config_code": "X123", "config_name": "32Q-TOR-T3",
                "data_type": "Demand", "ttl": 4,
                "ym": "2026-07", "week": "WK27",
                "date": "2026-07-06", "quantity": 100,
            },
            {
                "country": "爱尔兰", "category": "交换机整机",
                "config_code": "X123", "config_name": "32Q-TOR-T3",
                "data_type": "Supply", "ttl": 4,
                "ym": "2026-07", "week": "WK27",
                "date": "2026-07-06", "quantity": 120,
            },
        ]
        make_dsp_upload(
            db, vendor="Arista", item="X", sub_item="Y",
            version_date="2026-06-29", fact_rows=fact_rows,
        )
        make_week_dt(db, "2026-07-06", year_id=2026, month_id=7,
                     week_id=27, is_week_start=True)

        payload = {
            "pivot_type": "demand_plus_supply",
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_dates": ["2026-06-29"],
            "years": [2026],
        }
        resp = await client.post("/api/pivot-query/export", json=payload)
        assert resp.status_code == 200

        sheets = _parse_xlsx_all_sheets(resp.content)
        assert "透视结果" in sheets
        _, rows = sheets["透视结果"]
        # 每业务组 4 行
        data_types = [r[4] for r in rows]
        assert data_types == ["Demand", "Supply", "TTL_GAP", "Rolling_TTLGAP"]

    @pytest.mark.asyncio
    async def test_pivot_export_sheet2_snapshot(
        self, client, db, make_dsp_upload, make_week_dt,
    ):
        """sheet 2「查询参数快照」内容正确。"""
        make_dsp_upload(
            db, vendor="Arista", item="X", sub_item="Y",
            version_date="2026-06-29", fact_rows=[],
        )
        make_week_dt(db, "2026-07-06", year_id=2026, month_id=7,
                     week_id=27, is_week_start=True)

        payload = {
            "pivot_type": "demand",
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_dates": ["2026-06-29", "2026-07-15"],
            "years": [2026],
        }
        resp = await client.post("/api/pivot-query/export", json=payload)
        assert resp.status_code == 200

        sheets = _parse_xlsx_all_sheets(resp.content)
        assert "查询参数快照" in sheets
        headers, rows = sheets["查询参数快照"]
        assert headers == ["pivot_type", "vendor", "item", "sub_item", "version_dates"]
        assert len(rows) == 1
        assert rows[0][0] == "demand"
        assert rows[0][1] == "Arista"
        assert rows[0][2] == "X"
        assert rows[0][3] == "Y"
        # version_dates 用 "; " 拼接
        assert rows[0][4] == "2026-06-29; 2026-07-15"


# ---------- TC03~TC06: 错误约定 ----------


class TestPivotExportErrors:
    """错误约定：级联 / 笛卡尔积 / demand_plus_supply 多 version_date。"""

    @pytest.mark.asyncio
    async def test_pivot_export_422_cascade_validation(self, client):
        """422 级联：传 config_names 不传 categories。"""
        payload = {
            "pivot_type": "demand",
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_dates": ["2026-06-29"],
            "config_names": ["32Q-TOR-T3"],
            "years": [2026],
        }
        resp = await client.post("/api/pivot-query/export", json=payload)
        assert resp.status_code == 422

    @pytest.mark.asyncio
    async def test_pivot_export_422_cartesian_overflow(
        self, client, db, make_dsp_upload, make_week_dt, monkeypatch,
    ):
        """422 笛卡尔积超限：monkeypatch MAX_CARTESIAN=5。"""
        monkeypatch.setattr(pivot_query, "MAX_CARTESIAN", 5)

        fact_rows = []
        for i in range(10):
            fact_rows.append({
                "country": f"country_{i}",
                "category": "交换机整机",
                "config_code": f"X{i:03d}",
                "config_name": f"cfg_{i}",
                "data_type": "Demand",
                "ttl": 4,
                "ym": "2026-07",
                "week": "WK27",
                "date": "2026-07-06",
                "quantity": 1,
            })
        make_dsp_upload(
            db, vendor="Arista", item="X", sub_item="Y",
            version_date="2026-06-29", fact_rows=fact_rows,
        )
        make_week_dt(db, "2026-07-06", year_id=2026, month_id=7,
                     week_id=27, is_week_start=True)

        payload = {
            "pivot_type": "demand",
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_dates": ["2026-06-29"],
            "years": [2026],
        }
        resp = await client.post("/api/pivot-query/export", json=payload)
        assert resp.status_code == 422
        assert "cartesian product" in resp.json()["detail"]

    @pytest.mark.asyncio
    async def test_pivot_export_422_demand_plus_supply_multi_versions(
        self, client,
    ):
        """422 demand_plus_supply + 多 version_date。"""
        payload = {
            "pivot_type": "demand_plus_supply",
            "vendor": "Arista",
            "item": "X",
            "sub_item": "Y",
            "version_dates": ["2026-06-29", "2026-07-15"],
            "years": [2026],
        }
        resp = await client.post("/api/pivot-query/export", json=payload)
        assert resp.status_code == 422


# ---------- TC07: 空数据 ----------


class TestPivotExportEmpty:
    """row_groups=[] → sheet 1 仅表头，sheet 2 仍正确。"""

    @pytest.mark.asyncio
    async def test_pivot_export_200_empty_row_groups(
        self, client, db, make_dsp_upload, make_week_dt,
    ):
        # batch 不存在 → row_groups=[]
        make_week_dt(db, "2026-07-06", year_id=2026, month_id=7,
                     week_id=27, is_week_start=True)

        payload = {
            "pivot_type": "demand",
            "vendor": "NonExistent",
            "item": "X",
            "sub_item": "Y",
            "version_dates": ["2026-06-29"],
            "years": [2026],
        }
        resp = await client.post("/api/pivot-query/export", json=payload)
        assert resp.status_code == 200

        sheets = _parse_xlsx_all_sheets(resp.content)
        headers, rows = sheets["透视结果"]
        assert headers[:7] == [
            "国家", "类别", "配置代码", "配置名称",
            "数据类型", "TTL", "版本日期",
        ]
        # row_groups=[] 时，SQL JOIN 自然无输出 → period_columns 也为空
        # 断言：sheet 1 仅 7 列基础表头，无数据行
        assert len(headers) == 7
        assert rows == []

        # sheet 2 快照仍正确
        assert "查询参数快照" in sheets
        snap_headers, snap_rows = sheets["查询参数快照"]
        assert snap_headers == [
            "pivot_type", "vendor", "item", "sub_item", "version_dates",
        ]
        assert snap_rows[0][0] == "demand"
        assert snap_rows[0][1] == "NonExistent"