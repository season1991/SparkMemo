"""DSP Excel 解析模块（v0.5.3：列头文本匹配）。

本模块是纯函数模块，不依赖 SQLAlchemy / FastAPI；可被路由层直接调用，也可被单测直接测试。

**v0.5.3 变更**：解析不再依赖字面列号；改用 row 1 列头文本匹配。
- 静态字段（country / category / config_code / data_type / ttl）通过行 1 文本首匹配定位；
- `update_by` **不再作为周列起点边界**；周列起点由行 1 中匹配 `YYYY-MM` 模式的 cell 自动识别（一个段起点之后的连续 col 共享同一 ym，直到下一段起点）；
- 列头文本归一化规则：strip 首尾空白 → 去前缀 `*` → strip → 小写（`"  *Country  "` 与 `"country"` 等价）。

**关于隐藏列**：v0.5.3 **不**根据 `column_dimensions.hidden` 过滤列或 cell。
原因：实际 DSP 模板常把结构性 / 已废弃列（`*BU` / `Config Name` 等）设为 hidden 做分组视图，
   这些列不参与 `_HEADER_TARGETS` 匹配，跳不跳过都不影响解析；
   而在真实样本 `Arista-…xlsx` 中，关键列 `Category` (col E) 也被 hidden —— 若按 hidden 过滤，
   会导致该文件 422 列缺失。所以 hidden 视作 UI 状态不参与解析判定；
   若未来"按 hidden 跳列"成为业务需求，扩展点固定在 `_resolve_columns` / `_ym_segments` 内。

跳过规则（与 spec §跳过规则 一一对应）：
- 行级：R1（Country+ConfigCode 同时空）、R2（DataType 不是 Demand/Supply）、R3（无有效周列）；
- 列级：C1（行 2 周编号空）、C2（行 3 周起始日空）、C3（无 ym）、C4（quantity 空/None/0）。

异常：
- SheetMissingError → 422
- BadHeaderError → 422
- BadQuantityError → 400
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Optional

import openpyxl


SHEET_NAME = "DSP"

ROW_WEEK_NO = 2
ROW_WEEK_DATE = 3
ROW_DATA_START = 4

DATA_TYPES_KEPT = ("Demand", "Supply")

# 列头匹配目标；每个 key 是一组规范化别名，归一化后命中即停止遍历该列的所有 key。
# 关键列缺失任意一个 → BadHeaderError（v0.5.3）
# v0.5.5 新增 config_name（Config Name）列，可选列（缺失不报错，入库为 None）
_HEADER_TARGETS: dict[str, tuple[str, ...]] = {
    "country":     ("country",),
    "category":    ("category",),
    "config_code": ("config code", "configcode"),  # 容错：'Config Code' / 'ConfigCode' / 'configcode'
    "config_name": ("config name", "configname"),  # 可选列：'*Config Name' / 'Config Name' / 'ConfigName'
    "data_type":   ("data type", "datatype"),
    "ttl":         ("ttl",),
}

# 必填列：缺失任意一个 → BadHeaderError；config_name 为可选列，不在其中
_REQUIRED_COLUMNS = {"country", "category", "config_code", "data_type", "ttl"}

# YYYY-MM 模式：行 1 携带这种格式的 cell 视为 ym 段起点
_YM_PATTERN = re.compile(r"^\d{4}-\d{2}$")


class SheetMissingError(Exception):
    """Sheet 名不是 'DSP' → 路由层映射 422。"""


class BadHeaderError(Exception):
    """关键列（country/category/config_code/data_type/ttl）在行 1 缺失 → 路由层映射 422。
    config_name 为可选列，缺失时入库为 None，不报错。"""


class BadQuantityError(Exception):
    """quantity 含非数字字符串 / 非整数浮点 → 路由层映射 400。"""


@dataclass
class FactRow:
    """解析后的单条事实行。"""

    country: Optional[str]
    category: Optional[str]
    config_code: Optional[str]
    config_name: Optional[str]
    data_type: Optional[str]
    ttl: Optional[int]
    ym: str
    week: str
    date: str
    quantity: int


def parse_filename(filename: str) -> tuple[str, str, str]:
    """截掉扩展名后按 '-' 切分，返回前 3 段；段数 < 3 抛 ValueError。

    实现严格按 spec 给出的伪代码。

    **状态（v0.5.1 起）**：此函数**不再被** `POST /api/dsp-uploads` 调用——
    前端要求用户在 UI 内提供 vendor / item / sub_item，服务端仅校验与入库。
    本函数保留供以下场景使用：
    - 导入 / 迁移遗留脚本（spec 预留的「其他调用者」）；
    - 命令行调试工具；
    - 单元测试覆盖同名 spec §Test Plan §1 时的引用。
    后续如确认所有调用方都已迁移到 Form 字段，可安全删除本函数。
    """
    if not filename:
        raise ValueError("filename is required")
    stem = filename.rsplit(".", 1)[0] if "." in filename else filename
    parts = stem.split("-")
    if len(parts) < 3:
        raise ValueError("filename must contain at least 3 segments separated by '-'")
    return parts[0], parts[1], parts[2]


def _cell_str(value) -> str:
    """统一把 cell 值转字符串并 strip；None 返回空串。"""
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value) -> str:
    """列头文本归一化：strip → 去前缀 '*' → strip → lower。

    示例：
      "  *Country  "  → "country"
      "Data Type"     → "data type"
      None            → ""
    """
    if value is None:
        return ""
    s = str(value).strip()
    if s.startswith("*"):
        s = s[1:].strip()
    return s.lower()


def _resolve_columns(ws) -> dict[str, int]:
    """列头匹配：扫描 row 1 首匹配 `_HEADER_TARGETS` 中任一别名。

    v0.5.3 不根据 `column_dimensions.hidden` 过滤：详见模块顶部 docstring。
    返回: `{"country": col_idx, "category": ..., ...}`。
    任一关键 key 不在返回值中 → 调用方应 raise BadHeaderError。
    """
    found: dict[str, int] = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = _normalize_header(ws.cell(row=1, column=c).value)
        if not v:
            continue
        for key, alts in _HEADER_TARGETS.items():
            if key in found:
                continue
            if v in alts:
                found[key] = c
                break
    return found


def _ym_segments(ws, max_col: int) -> dict[int, str]:
    """v0.5.3 替代旧的「col 13 起做前向传播」。

    扫描 row 1 全 cell：值匹配 `YYYY-MM` 即视为段起点，记录到该 col；
    该段起点之后的所有 col 都继承该 ym，直到下一个段起点。
    不考虑 hidden —— 见 `_resolve_columns` 注释。
    返回: `col → "YYYY-MM"`。
    """
    result: dict[int, str] = {}
    current = ""
    for c in range(1, max_col + 1):
        v = _cell_str(ws.cell(row=1, column=c).value)
        if v and _YM_PATTERN.match(v):
            current = v
        if current:
            result[c] = current
    return result


def _parse_ttl(value) -> Optional[int]:
    """TTL 容错：整数原样；浮点整数化；其它入 None（不阻断上传）。"""
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(s)
        except ValueError:
            try:
                f = float(s)
            except ValueError:
                return None
            return int(f) if f.is_integer() else None
    return None


def _parse_quantity(value, *, row: int, col: int) -> Optional[int]:
    """解析 quantity；返回 None 表示跳过该 cell；抛 BadQuantityError 表示阻断。

    规则（C4）：
    - None / 空字符串 → 跳过
    - bool → 跳过
    - 整数 0 / 浮点 0.0 → 跳过
    - 整数 >0 / 浮点整数 >0 → 入库 int
    - 浮点非整数（如 1.5）→ 阻断
    - 字符串：strip 后按 float 解析；非数字 → 阻断；整数化 → 入库；非整数浮点 → 阻断
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value != 0 else None
    if isinstance(value, float):
        if value == 0:
            return None
        if not value.is_integer():
            raise BadQuantityError(
                f"row {row} col {col}: quantity {value!r} is non-integer"
            )
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            f = float(s)
        except ValueError as exc:
            raise BadQuantityError(
                f"row {row} col {col}: quantity {value!r} is non-numeric"
            ) from exc
        if f == 0:
            return None
        if not f.is_integer():
            raise BadQuantityError(
                f"row {row} col {col}: quantity {value!r} is non-integer"
            )
        return int(f)
    raise BadQuantityError(
        f"row {row} col {col}: quantity {value!r} is unsupported type"
    )


def parse_excel(content: bytes) -> list[FactRow]:
    """解析 DSP Excel 字节流，返回展开后的事实行列表。

    v0.5.3 行为：
    1. 校验 sheet 名 = 'DSP'，否则 `SheetMissingError`；
    2. 在 row 1 列头文本匹配关键列（country/category/config_code/data_type/ttl），
       任一缺失 → `BadHeaderError`；
    3. 在 row 1 扫 `YYYY-MM` 段起点（跳过隐藏列），构建周列集合；
    4. 对 row 4+ 数据行按 R1/R2 行级过滤；对每个有效周列按 C1/C2/C3/C4/C5 列级过滤；
    5. 展开后返回 `[FactRow, ...]`。

    异常:
        SheetMissingError → 422
        BadHeaderError   → 422
        BadQuantityError → 400
    """
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise SheetMissingError(f"sheet '{SHEET_NAME}' not found")
    ws = wb[SHEET_NAME]

    try:
        # 1) 列头匹配（跳过隐藏列）
        cols = _resolve_columns(ws)
        for required in _REQUIRED_COLUMNS:
            if required not in cols:
                wb.close()
                raise BadHeaderError(
                    f"Excel header missing required column '{required}'"
                )

        # 2) ym 段起点扫描（row 1 全列匹配 YYYY-MM，隐藏列跳过）
        ym_at_col = _ym_segments(ws, ws.max_column)

        # 3) 有效周列：ym 存在 ∧ 行 2 周编号非空 ∧ 行 3 周起始日非空（隐藏列已在前两步排除）
        week_cols: list[tuple[int, str, str]] = []
        for c in sorted(ym_at_col.keys()):
            week_v = _cell_str(ws.cell(row=ROW_WEEK_NO, column=c).value)
            date_v = _cell_str(ws.cell(row=ROW_WEEK_DATE, column=c).value)
            if not week_v or not date_v:
                continue
            week_cols.append((c, week_v, date_v))

        facts: list[FactRow] = []

        # R3：如果完全没有有效周列，事实行集合必为空
        if not week_cols:
            return facts

        for r in range(ROW_DATA_START, ws.max_row + 1):
            country = _cell_str(ws.cell(row=r, column=cols["country"]).value)
            config_code = _cell_str(ws.cell(row=r, column=cols["config_code"]).value)

            # R1：Country 与 Config Code 同时为空 → 整行跳过
            if not country and not config_code:
                continue

            data_type = _cell_str(ws.cell(row=r, column=cols["data_type"]).value)
            # R2：Data Type 严格匹配 Demand/Supply
            if data_type not in DATA_TYPES_KEPT:
                continue

            category = _cell_str(ws.cell(row=r, column=cols["category"]).value) or None
            # config_name 为可选列（v0.5.5），缺失时入库为 None
            config_name_col = cols.get("config_name")
            config_name = _cell_str(ws.cell(row=r, column=config_name_col).value) or None if config_name_col else None
            ttl = _parse_ttl(ws.cell(row=r, column=cols["ttl"]).value)

            for c, week, date in week_cols:
                # C5（隐藏列）已在 _ym_segments 中排除；此处直接读即可
                q_raw = ws.cell(row=r, column=c).value
                # C4：数量解析
                quantity = _parse_quantity(q_raw, row=r, col=c)
                if quantity is None:
                    continue
                facts.append(
                    FactRow(
                        country=country or None,
                        category=category,
                        config_code=config_code or None,
                        config_name=config_name,
                        data_type=data_type,
                        ttl=ttl,
                        ym=ym_at_col[c],
                        week=week,
                        date=date,
                        quantity=quantity,
                    )
                )
        return facts
    finally:
        wb.close()
