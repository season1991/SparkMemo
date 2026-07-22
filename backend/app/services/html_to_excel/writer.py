"""JSON 中间表示 → xlsx 写入（SPEC §4.5）。

- 复用既有 `excel_export.py` 的 `_sanitize_formula` / `_auto_width`；
- 不引入 pandas；用 openpyxl 直接构建 workbook；
- 列类型映射见 SPEC §4.5.2；
- 输出文件后写入磁盘。
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..excel_export import _auto_width, _export_timestamp, _sanitize_formula
from .schemas import ExtractedCell, ExtractedControl


_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="EEEEEE")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_LINK_FORMAT = "@"
_WRAP_ALIGN = Alignment(wrap_text=True)


_NUMBER_FORMAT_MAP: dict[str, str] = {
    "number": "#,##0.00",
    "integer": "0",
    "date": "m/d/yyyy",
    "datetime": "m/d/yyyy h:mm am/pm",
    "boolean": "@",
    "text": "@",
    "status": "@",
    "tooltip": "@",
    "link": "@",
    "html": "@",
}


# Sheet name 在 Excel 限 31 字符；过长截断并去重
_MAX_SHEET_NAME = 31


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """生成可用的 sheet 名称（≤31 字符）。"""
    base = (name or "Sheet").strip()[:_MAX_SHEET_NAME] or "Sheet"
    base = base.replace("/", "_").replace("\\", "_").replace("?", "").replace("*", "").replace("[", "(").replace("]", ")")
    candidate = base
    suffix = 1
    while candidate in used or candidate == "":
        suffix += 1
        suffix_str = f"~{suffix}"
        candidate = (base[: _MAX_SHEET_NAME - len(suffix_str)] + suffix_str)
    used.add(candidate)
    return candidate


def _coerce_value(cell: ExtractedCell) -> tuple[Any, str]:
    """根据 cell.type 转成 openpyxl 可写入的值 + number_format。

    返回 `(value, number_format)`。
    """
    fmt = _NUMBER_FORMAT_MAP.get(cell.type, "@")

    if cell.type == "boolean":
        v = bool(cell.value) if cell.value is not None else False
        return v, fmt

    if cell.type == "number" or cell.type == "integer":
        if cell.value is None or cell.value == "":
            return "", fmt
        try:
            return float(cell.value) if cell.type == "number" else int(cell.value), fmt
        except (TypeError, ValueError):
            return str(cell.value), fmt

    if cell.type == "date":
        if cell.value is None or cell.value == "":
            return "", fmt
        # 接受 string 形式 M/d/yyyy
        parsed = _parse_date_only(str(cell.value))
        return (parsed or str(cell.value)), fmt

    if cell.type == "datetime":
        if cell.value is None or cell.value == "":
            return "", fmt
        parsed = _parse_datetime(str(cell.value))
        return (parsed or str(cell.value)), fmt

    if cell.type == "link":
        # 合成为一个字符串：`text (href)`，避免超链接
        text = cell.value if cell.value is not None else ""
        href = cell.href or ""
        if href and text and href not in text:
            return _sanitize_formula(f"{text} ({href})"), fmt
        return _sanitize_formula(text or href), fmt

    # text / status / tooltip / html / fallback
    sval = "" if cell.value is None else str(cell.value)
    return _sanitize_formula(sval), fmt


def _parse_date_only(text: str) -> datetime | None:
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _parse_datetime(text: str) -> datetime | None:
    for fmt in ("%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    # fallback：日期 + 时间灵活匹配
    return _parse_date_only(text)


class ExcelWriter:
    """把 `ExtractedControl` 转成 xlsx 并落盘。"""

    def write(self, control: ExtractedControl, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        # 默认 sheet 1 名 = matched_text；超长做安全处理
        used: set[str] = set()
        sheet_name = _safe_sheet_name(control.matched_text or control.title or "Sheet1", used)
        ws = wb.active
        ws.title = sheet_name

        # 表头
        headers = [c.key for c in control.columns]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=_sanitize_formula(h))
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
        ws.freeze_panes = "A2"

        # 数据
        for r_idx, row in enumerate(control.rows, start=2):
            for c_idx in range(len(control.columns)):
                cell_obj = row.cells[c_idx] if c_idx < len(row.cells) else ExtractedCell()
                value, fmt = _coerce_value(cell_obj)
                cell = ws.cell(row=r_idx, column=c_idx + 1, value=value)
                cell.number_format = fmt
                cell.alignment = _WRAP_ALIGN

        # 列宽自适应（基于所有 cell 长度的 visual 长度）
        all_rows: list[list[str]] = [[]]
        for row in control.rows:
            line: list[str] = []
            for c_idx in range(len(control.columns)):
                cell_obj = row.cells[c_idx] if c_idx < len(row.cells) else ExtractedCell()
                value, _ = _coerce_value(cell_obj)
                line.append("" if value is None else str(value))
            all_rows.append(line)
        widths = _auto_width(headers, all_rows)
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        wb.save(output_path)
        return output_path


def default_filename(matched_title: str, hint: str | None = None) -> str:
    """生成 `hint_{timestamp}.xlsx` 或 `{matched_title}_{timestamp}.xlsx`。"""
    ts = _export_timestamp()
    base = (hint or matched_title or "output").strip() or "output"
    # 文件名清洗（替换非法字符）
    safe = "".join(c if c.isalnum() or c in ("_", "-", ".") else "_" for c in base)
    return f"{safe}_{ts}.xlsx"
