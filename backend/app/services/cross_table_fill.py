"""跨表数据填充 Service 层（v0.6.0）。

包含三个主要函数：
- `parse_table(content, role)`：从 xlsx 字节流解析第一行表头与数据行，返回 (headers, rows)
- `execute_match(...)`：按主键归一化索引 + 遍历 target 应用 mappings，返回 ExecuteResult
- `build_xlsx(headers, rows)`：基于 openpyxl 生成结果 xlsx 二进制

异常类：
- `EmptyHeadersError` → 422
- `DuplicateHeadersError` → 422
- `BadCellTypeError` → 400
- `NoSheetError` → 422
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Optional

import openpyxl
from openpyxl import Workbook


# preview 返回上限
PREVIEW_LIMIT = 1000


class NoSheetError(Exception):
    """工作簿没有任何 sheet → 路由层映射 422。"""


class EmptyHeadersError(Exception):
    """表头去重后为空 → 422。`role` 表示是哪一张表（'target' / 'base'）。"""

    def __init__(self, role: str):
        self.role = role
        super().__init__(f"{role}_headers is empty")


class DuplicateHeadersError(Exception):
    """表头含重复字段名 → 422。"""

    def __init__(self, role: str, name: str):
        self.role = role
        self.name = name
        super().__init__(f"{role}_headers contains duplicate: '{name}'")


class BadCellTypeError(Exception):
    """数据行 cell 含不支持的类型（如 datetime） → 400。"""

    def __init__(self, row_index: int, col_name: str, type_name: str):
        self.row_index = row_index
        self.col_name = col_name
        self.type_name = type_name
        super().__init__(
            f"cell at row {row_index} col '{col_name}': unsupported type {type_name}"
        )


# ============================================================
# 解析
# ============================================================


def _normalize_cell_value(raw: Any, row_index: int, col_name: str) -> Any:
    """将 cell 原值归一化为原生 Python 类型；datetime 等不支持类型抛 BadCellTypeError。"""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)):
        return raw
    if isinstance(raw, str):
        return raw.strip()
    raise BadCellTypeError(row_index, col_name, type(raw).__name__)


def parse_table(content: bytes, role: str) -> tuple[list[str], list[dict]]:
    """解析 xlsx 字节流，返回 (headers, rows)。

    参数:
        content: 原始 xlsx 文件字节流。
        role: 'target' 或 'base'，仅用于错误信息构造。

    返回:
        tuple[list[str], list[dict]]: (表头列表, 数据行列表);
        数据行每项是 dict[str, Any]，key 与表头一一对应。

    异常:
        NoSheetError: 工作簿无任何 sheet。
        EmptyHeadersError: 去重后表头为空。
        DuplicateHeadersError: 表头含重复字段名。
        BadCellTypeError: 数据行含不支持类型 cell。
    """
    wb = openpyxl.load_workbook(BytesIO(content), read_only=False)
    if not wb.sheetnames:
        raise NoSheetError("workbook has no sheet")
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    headers_raw = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    wb.close()

    # headers：strip；空字符串保留；None 视为空
    headers: list[str] = []
    for h in headers_raw:
        if h is None:
            headers.append("")
        else:
            headers.append(str(h).strip())

    # 重复校验（空白 header 不计入）
    seen: set[str] = set()
    for h in headers:
        if not h:
            continue
        if h in seen:
            raise DuplicateHeadersError(role, h)
        seen.add(h)

    non_empty = [h for h in headers if h]
    if not non_empty:
        raise EmptyHeadersError(role)

    # 读数据行
    rows: list[dict] = []
    wb = openpyxl.load_workbook(BytesIO(content), read_only=False)
    ws = wb[sheet_name]
    for r_idx in range(2, ws.max_row + 1):
        any_present = False
        values: list[Any] = []
        for c_idx, h in enumerate(headers, start=1):
            if not h:
                values.append(None)
                continue
            v = ws.cell(row=r_idx, column=c_idx).value
            v_normalized = _normalize_cell_value(v, r_idx, h)
            if v_normalized is not None and v_normalized != "":
                any_present = True
            values.append(v_normalized)
        if not any_present:
            continue
        row_dict: dict[str, Any] = {}
        for h, v in zip(headers, values):
            if h:
                row_dict[h] = v
        rows.append(row_dict)
    wb.close()

    return non_empty, rows


# ============================================================
# 匹配
# ============================================================


@dataclass
class MappingSpec:
    """单条映射规则（来自 Pydantic schema 的 Python 表示）。"""

    base_field: str
    target_field: str
    mode: str  # 'overwrite' / 'new_column'


@dataclass
class ExecuteConfig:
    """execute 入参（Python 表示，不依赖 ORM）。"""

    target_keys: list[str]
    base_keys: list[str]
    mappings: list[MappingSpec]
    join_mode: str
    match_mode: str
    case_sensitive: bool
    trim_strings: bool


@dataclass
class ExecuteResult:
    """execute 输出。"""

    final_headers: list[str]
    final_rows: list[list[Any]]
    result_row_count: int
    filled_count: int
    unmatched_count: int
    multi_match_count: int


def _normalize(value: Any, cfg: ExecuteConfig) -> Optional[str]:
    """主键值归一化（spec §匹配算法 §2）。"""
    if value is None:
        return None
    s = str(value)
    if cfg.trim_strings:
        s = s.strip()
    if not cfg.case_sensitive:
        s = s.lower()
    return s


def _to_str(v: Any) -> str:
    """将任意值转为字符串（merge_multi 拼接用）。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return str(v)


def execute_match(
    *,
    target_headers: list[str],
    target_rows: list[dict],
    base_headers: list[str],
    base_rows: list[dict],
    cfg: ExecuteConfig,
) -> ExecuteResult:
    """按主键匹配 + 应用 mappings，返回 ExecuteResult。

    算法详见 spec §匹配算法 §3。
    """
    # 1. base 索引
    base_index: dict[tuple, list[dict]] = defaultdict(list)
    for base_row in base_rows:
        raw_key = tuple(base_row.get(k) for k in cfg.base_keys)
        norm_key = tuple(_normalize(v, cfg) for v in raw_key)
        if any(v is None or v == "" for v in norm_key):
            continue
        base_index[norm_key].append(base_row)

    # 2. 计算 final_headers（处理 new_column 的 _filled 后缀）
    final_headers = list(target_headers)
    added_new_columns: list[str] = []
    existing_set = set(final_headers)

    # mapping order → new_column 实际最终列名
    mapping_to_final_col: list[Optional[str]] = []
    for m in cfg.mappings:
        if m.mode == "new_column":
            desired = m.target_field
            if desired not in existing_set and desired not in added_new_columns:
                final_col = desired
            else:
                base_name = desired + "_filled"
                candidate = base_name
                n = 2
                while candidate in existing_set or candidate in added_new_columns:
                    candidate = f"{base_name}_{n}"
                    n += 1
                final_col = candidate
            added_new_columns.append(final_col)
            mapping_to_final_col.append(final_col)
        else:
            # overwrite 走 target 原列
            mapping_to_final_col.append(m.target_field)

    final_headers = final_headers + added_new_columns
    # final_headers 的列名 → index
    final_col_index = {h: i for i, h in enumerate(final_headers)}

    # 3. 遍历 target
    final_rows: list[list[Any]] = []
    filled_count = 0
    unmatched_count = 0
    multi_match_count = 0

    for target_row in target_rows:
        raw_key = tuple(target_row.get(k) for k in cfg.target_keys)
        norm_key = tuple(_normalize(v, cfg) for v in raw_key)

        out_row: list[Any] = [target_row.get(h) for h in target_headers]
        out_row += [None] * (len(final_headers) - len(target_headers))

        # 主键任一为空 → unmatched
        if any(v is None or v == "" for v in norm_key):
            final_rows.append(out_row)
            unmatched_count += 1
            continue

        candidates = base_index.get(norm_key, [])
        if not candidates:
            final_rows.append(out_row)
            unmatched_count += 1
            continue

        if len(candidates) > 1:
            multi_match_count += 1

        # 选 candidates
        if cfg.match_mode == "first":
            chosen_list = [candidates[0]]
        elif cfg.match_mode == "last":
            chosen_list = [candidates[-1]]
        else:
            chosen_list = candidates

        # 应用 mappings
        n_filled_here = 0
        for m, final_col in zip(cfg.mappings, mapping_to_final_col):
            vals: list[Any] = []
            for c in chosen_list:
                cv = c.get(m.base_field)
                if cv is None or cv == "":
                    continue
                vals.append(cv)
            if not vals:
                continue
            if cfg.match_mode == "merge_multi" and len(vals) > 1:
                fill_value = ";".join(_to_str(v) for v in vals)
            else:
                fill_value = vals[0]
            out_row[final_col_index[final_col]] = fill_value
            n_filled_here += 1

        if n_filled_here > 0:
            filled_count += 1
        final_rows.append(out_row)

    # 4. inner join：剔除 unmatched 行
    if cfg.join_mode == "inner":
        keep: list[list[Any]] = []
        for target_row, out_row in zip(target_rows, final_rows):
            raw_key = tuple(target_row.get(k) for k in cfg.target_keys)
            norm_key = tuple(_normalize(v, cfg) for v in raw_key)
            if any(v is None or v == "" for v in norm_key):
                continue
            if norm_key in base_index:
                keep.append(out_row)
        final_rows = keep
        unmatched_count = 0

    return ExecuteResult(
        final_headers=final_headers,
        final_rows=final_rows,
        result_row_count=len(final_rows),
        filled_count=filled_count,
        unmatched_count=unmatched_count,
        multi_match_count=multi_match_count,
    )


# ============================================================
# 输出 xlsx
# ============================================================


def build_xlsx(headers: list[str], rows: list[list[Any]]) -> bytes:
    """生成 xlsx 二进制流。

    cell 写入规则：
    - None → 空白
    - 数字（int / float，不含 bool）→ 数字 cell
    - 其它 → str() 后写入字符串 cell
    """
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append([_to_xlsx_cell(c) for c in r])
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def _to_xlsx_cell(v: Any) -> Any:
    """openpyxl 写入 cell 时的归一化。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return v
    return str(v)
